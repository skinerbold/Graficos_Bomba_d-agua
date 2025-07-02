import sys
import numpy as np
import traceback
from scipy.interpolate import interp1d
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
import openpyxl.styles
import openpyxl.utils
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                            QLabel, QPushButton, QFileDialog, QLineEdit, QMessageBox,
                            QGroupBox, QScrollArea, QInputDialog, QDialog, QDialogButtonBox,
                            QTabWidget, QTableWidget, QTableWidgetItem, QAbstractItemView,
                            QHeaderView, QComboBox, QCheckBox)
from PyQt5.QtGui import QPixmap, QImage, QPainter, QPen, QColor, QPainterPath, QDoubleValidator
from PyQt5.QtCore import Qt, QPoint, QRect

class ImageWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self.image = QImage()
        self.drawing_scale = False
        self.drawing_points = False
        self.scale_rect = QRect()
        self.start_point = QPoint()
        self.points = []
        self.rotor_points = {}
        self.rotor_rpm = {}  # Dicionário para armazenar RPM de cada rotor
        self.current_rotor = None
        self.scale_values = {}

    def reset_data(self):
        self.drawing_scale = False
        self.drawing_points = False
        self.scale_rect = QRect()
        self.start_point = QPoint()
        self.points = []
        self.rotor_points = {}
        self.rotor_rpm = {}  # Resetar também os RPMs
        self.current_rotor = None
        self.scale_values = {}

    def load_image(self, path):
        self.image = QImage(path)
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        if not self.image.isNull():
            img = self.image.scaled(self.size(), Qt.KeepAspectRatio)
            painter.drawImage(0, 0, img)

        if self.drawing_scale and not self.scale_rect.isNull():
            painter.setPen(QPen(Qt.red, 2, Qt.DashLine))
            painter.drawRect(self.scale_rect)

        for rotor, points in self.rotor_points.items():
            color = QColor(*self.get_color_for_rotor(rotor))
            painter.setPen(QPen(color, 8))
            for point in points:
                painter.drawEllipse(point['pos'], 3, 3)

    def get_color_for_rotor(self, rotor):
        rotors = list(self.rotor_points.keys())
        index = rotors.index(rotor) % 5
        return [
            (255, 0, 0), (0, 255, 0), (0, 0, 255),
            (255, 255, 0), (255, 0, 255)
        ][index]

    def mousePressEvent(self, event):
        if self.drawing_scale:
            self.start_point = event.pos()
            self.scale_rect = QRect(self.start_point, self.start_point)
        elif self.drawing_points:
            self.add_efficiency_point(event.pos())

    def mouseMoveEvent(self, event):
        if self.drawing_scale and not self.start_point.isNull():
            self.scale_rect = QRect(self.start_point, event.pos()).normalized()
            self.update()

    def mouseReleaseEvent(self, event):
        if self.drawing_scale:
            self.drawing_scale = False
            if self.main_window:
                self.main_window.update_scale_inputs(self.scale_rect)

    def add_efficiency_point(self, point):
        if not self.current_rotor:
            QMessageBox.warning(self, "Erro", "Selecione um rotor primeiro!")
            return

        if not self.scale_rect.contains(point):
            QMessageBox.warning(self, "Erro", "Ponto fora da área de escala!")
            return

        efficiency, ok = QInputDialog.getDouble(
            self, "Eficiência", f"Digite a eficiência para o rotor {self.current_rotor} (%):",
            min=0, max=100, decimals=1
        )
        
        if ok:
            if self.current_rotor not in self.rotor_points:
                self.rotor_points[self.current_rotor] = []
            
            self.rotor_points[self.current_rotor].append({
                'pos': point,
                'efficiency': efficiency
            })
            self.update()

class PumpAnalyzer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Análise de Curvas de Bomba")
        self.setGeometry(100, 100, 1200, 800)
        self.scale_values = {'x0': 0, 'y0': 0, 'x1': 0, 'y1': 0}
        self.manual_system_points = None
        self.direct_equation_params = None
        # Adicionar variáveis para a segunda curva do sistema
        self.manual_system_points_2 = None
        self.direct_equation_params_2 = None
        self.setup_ui()
        
    def setup_ui(self):
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)

        self.image_widget = ImageWidget(self)
        self.image_widget.scale_values = self.scale_values
        
        control_panel = QScrollArea()
        control_content = QWidget()
        control_layout = QVBoxLayout(control_content)

        # Grupo de arquivo
        file_group = QGroupBox("Arquivo")
        file_layout = QVBoxLayout()
        btn_load = QPushButton("Carregar Imagem")
        btn_load.clicked.connect(self.load_image)
        file_layout.addWidget(btn_load)
        file_group.setLayout(file_layout)
        control_layout.addWidget(file_group)

        # Grupo de escala
        scale_group = QGroupBox("Configuração de Escala")
        scale_layout = QVBoxLayout()
        self.scale_inputs = {}
        for axis in ['x0', 'y0', 'x1', 'y1']:
            layout = QHBoxLayout()
            label = QLabel(f"{axis}:")
            inp = QLineEdit()
            inp.setPlaceholderText("Valor real")
            layout.addWidget(label)
            layout.addWidget(inp)
            scale_layout.addLayout(layout)
            self.scale_inputs[axis] = inp
        
        btn_set_scale = QPushButton("Definir Área do Gráfico")
        btn_set_scale.clicked.connect(self.start_scale_selection)
        scale_layout.addWidget(btn_set_scale)
        
        btn_save_scale = QPushButton("Salvar Escala")
        btn_save_scale.clicked.connect(self.save_scale)
        scale_layout.addWidget(btn_save_scale)
        scale_group.setLayout(scale_layout)
        control_layout.addWidget(scale_group)

        # Grupo de rotores
        rotor_group = QGroupBox("Gerenciamento de Rotores")
        rotor_layout = QVBoxLayout()
        self.rotor_input = QLineEdit()
        self.rotor_input.setPlaceholderText("Diâmetro do rotor (mm)")
        rotor_layout.addWidget(self.rotor_input)
        
        btn_add_rotor = QPushButton("Adicionar Rotor")
        btn_add_rotor.clicked.connect(self.add_rotor)
        rotor_layout.addWidget(btn_add_rotor)
        
        # Adicionar seletor de rotores
        rotor_selector_layout = QHBoxLayout()
        rotor_selector_layout.addWidget(QLabel("Rotor atual:"))
        self.rotor_selector = QComboBox()
        self.rotor_selector.currentTextChanged.connect(self.select_current_rotor)
        rotor_selector_layout.addWidget(self.rotor_selector)
        rotor_layout.addLayout(rotor_selector_layout)
        
        # Adicionar botão para alterar RPM
        btn_change_rpm = QPushButton("Alterar RPM do Rotor")
        btn_change_rpm.clicked.connect(self.change_rotor_rpm)
        rotor_layout.addWidget(btn_change_rpm)
        
        btn_select_points = QPushButton("Selecionar Pontos")
        btn_select_points.clicked.connect(self.start_point_selection)
        rotor_layout.addWidget(btn_select_points)
        rotor_group.setLayout(rotor_layout)
        control_layout.addWidget(rotor_group)

        # Grupo de exportação
        export_group = QGroupBox("Exportação")
        export_layout = QVBoxLayout()
        
        # Curva do Sistema 1
        export_layout.addWidget(QLabel("Curva do Sistema 1:"))
        self.system_curve_mode = QComboBox()
        self.system_curve_mode.addItems(["Pontos Manual", "Equação Direta"])
        btn_set_curve = QPushButton("Configurar Curva do Sistema 1")
        btn_set_curve.clicked.connect(self.configure_system_curve)
        
        export_layout.addWidget(QLabel("Método da Curva do Sistema 1:"))
        export_layout.addWidget(self.system_curve_mode)
        export_layout.addWidget(btn_set_curve)
        
        # Curva do Sistema 2
        export_layout.addWidget(QLabel("Curva do Sistema 2:"))
        self.system_curve_mode_2 = QComboBox()
        self.system_curve_mode_2.addItems(["Nenhuma", "Pontos Manual", "Equação Direta"])
        btn_set_curve_2 = QPushButton("Configurar Curva do Sistema 2")
        btn_set_curve_2.clicked.connect(self.configure_system_curve_2)
        
        export_layout.addWidget(QLabel("Método da Curva do Sistema 2:"))
        export_layout.addWidget(self.system_curve_mode_2)
        export_layout.addWidget(btn_set_curve_2)
        
        btn_export = QPushButton("Gerar Relatório Excel")
        btn_export.clicked.connect(self.export_to_excel)
        export_layout.addWidget(btn_export)
        
        export_group.setLayout(export_layout)
        control_layout.addWidget(export_group)

        control_panel.setWidget(control_content)
        main_layout.addWidget(self.image_widget, 70)
        main_layout.addWidget(control_panel, 30)
        self.setCentralWidget(main_widget)

    def load_image(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Selecionar Imagem", "", "Imagens (*.png *.jpg *.jpeg *.bmp)"
        )
        if path:
            self.image_widget.reset_data()
            self.image_widget.load_image(path)
            QMessageBox.information(self, "Sucesso", "Imagem carregada com sucesso!")

    def start_scale_selection(self):
        if self.image_widget.image.isNull():
            QMessageBox.warning(self, "Erro", "Carregue uma imagem primeiro!")
            return
        self.image_widget.drawing_scale = True
        QMessageBox.information(self, "Instruções",            "Selecione a área do gráfico arrastando o mouse da origem (x0,y0) até o extremo (x1,y1)")

    def update_scale_inputs(self, rect):
        self.image_widget.scale_rect = rect
        QMessageBox.information(self, "Área Selecionada", 
            "Área do gráfico definida! Agora insira os valores reais para x0, y0, x1, y1 e clique em 'Salvar Escala'.")
        self.scale_inputs['x0'].setText("0")
        self.scale_inputs['y0'].setText("0")
        self.scale_inputs['x1'].setText(str(rect.width()))
        self.scale_inputs['y1'].setText(str(rect.height()))
        self.save_scale()

    def save_scale(self):
        try:
            for axis in self.scale_inputs:
                self.scale_values[axis] = self.convert_br_float(self.scale_inputs[axis].text())
            QMessageBox.information(self, "Sucesso", "Escala configurada!")
        except ValueError:
            QMessageBox.critical(self, "Erro", "Valores de escala inválidos")

    def add_rotor(self):
        rotor_name, ok = QInputDialog.getText(self, "Novo Rotor", "Nome do rotor:")
        if ok and rotor_name:
            # Solicitar RPM do rotor
            rpm, rpm_ok = QInputDialog.getDouble(
                self, "RPM do Rotor", f"Digite o RPM do rotor {rotor_name}:",
                value=1750, min=1, max=10000, decimals=0
            )
            
            if rpm_ok:
                self.image_widget.rotor_points[rotor_name] = []
                self.image_widget.rotor_rpm[rotor_name] = rpm
                self.update_rotor_list()
                self.image_widget.current_rotor = rotor_name
                self.rotor_selector.setCurrentText(rotor_name)
                QMessageBox.information(self, "Sucesso", f"Rotor {rotor_name} adicionado com {rpm} RPM")
    
    def update_rotor_list(self):
        """Atualiza a lista de rotores no seletor"""
        if hasattr(self, 'rotor_selector'):
            current_text = self.rotor_selector.currentText()
            self.rotor_selector.clear()
            
            rotors = list(self.image_widget.rotor_points.keys())
            self.rotor_selector.addItems(rotors)
            
            # Manter seleção atual se ainda existir
            if current_text in rotors:
                self.rotor_selector.setCurrentText(current_text)
            elif rotors:
                self.rotor_selector.setCurrentText(rotors[0])
                self.image_widget.current_rotor = rotors[0]

    def select_current_rotor(self, rotor_name):
        """Seleciona o rotor atual"""
        if rotor_name and rotor_name in self.image_widget.rotor_points:
            self.image_widget.current_rotor = rotor_name
            self.image_widget.update()
    
    def change_rotor_rpm(self):
        """Permite alterar o RPM de um rotor existente"""
        if not self.image_widget.rotor_points:
            QMessageBox.warning(self, "Erro", "Nenhum rotor disponível!")
            return
        
        # Selecionar rotor
        rotors = list(self.image_widget.rotor_points.keys())
        rotor, ok = QInputDialog.getItem(
            self, "Selecionar Rotor", "Escolha o rotor:", rotors, 0, False
        )
        
        if not ok:
            return
        
        # Obter RPM original
        original_rpm = self.image_widget.rotor_rpm.get(rotor, 1750)
        
        # Solicitar novo RPM
        new_rpm, ok = QInputDialog.getDouble(
            self, "Novo RPM", f"RPM atual: {original_rpm}\nDigite o novo RPM:",
            value=original_rpm, min=1, max=10000, decimals=0
        )
        
        if not ok or new_rpm <= 0:
            return
        
        # Calcular razão de RPM
        rpm_ratio = new_rpm / original_rpm
        
        # Criar novo nome do rotor
        new_rotor_name = f"{rotor}_{int(new_rpm)}RPM"
        
        # Verificar se o novo nome já existe
        if new_rotor_name in self.image_widget.rotor_points:
            QMessageBox.warning(self, "Aviso", f"Já existe um rotor com RPM {new_rpm}!")
            return
        
        # Aplicar leis de afinidade aos pontos
        original_points = self.image_widget.rotor_points[rotor]
        new_points = []
        
        for point in original_points:
            # Obter valores originais
            original_pos = point['pos']
            original_x = self.convert_to_real(original_pos, 'x')  # Vazão
            original_y = self.convert_to_real(original_pos, 'y')  # Altura
            original_eff = point['efficiency']
            
            # Q2/Q1 = N2/N1
            new_flow = original_x * rpm_ratio
            # H2/H1 = (N2/N1)²
            new_head = original_y * (rpm_ratio ** 2)
            # Eficiência permanece constante
            new_efficiency = original_eff
            
            # Converter de volta para coordenadas de pixel
            new_pos = self.convert_from_real_point(new_flow, new_head)
            
            new_points.append({
                'pos': new_pos,
                'efficiency': new_efficiency
            })
        
        # Adicionar novo rotor
        self.image_widget.rotor_points[new_rotor_name] = new_points
        self.image_widget.rotor_rpm[new_rotor_name] = new_rpm
        
        # Atualizar interface
        if hasattr(self, 'rotor_selector'):
            self.rotor_selector.addItem(new_rotor_name)
            self.rotor_selector.setCurrentText(new_rotor_name)
        
        self.image_widget.current_rotor = new_rotor_name
        self.image_widget.update()
        
        QMessageBox.information(
            self, "Sucesso", 
            f"Rotor '{new_rotor_name}' criado com RPM {new_rpm}!"
        )
            
    def change_rotor_rpm_manual(self, rotor, parent_dialog):
        # Verificar se o rotor selecionado é válido
        if not rotor or rotor not in self.image_widget.rotor_points:
            QMessageBox.warning(self, "Aviso", "Selecione um rotor válido!")
            return
            
        # Criar diálogo para o novo RPM
        rpm_dialog = QDialog(parent_dialog)
        rpm_dialog.setWindowTitle("Alterar RPM do Rotor")
        rpm_layout = QVBoxLayout()
        
        # Entrada do novo RPM
        rpm_layout.addWidget(QLabel(f"Digite o novo RPM para o rotor {rotor}:"))
        rpm_input = QLineEdit()
        rpm_input.setValidator(QDoubleValidator(1, 10000, 0))
        
        # Se o rotor já tem um RPM definido, mostrar como valor padrão
        if rotor in self.image_widget.rotor_rpm:
            rpm_input.setText(str(self.image_widget.rotor_rpm[rotor]))
        else:
            rpm_input.setText("1750")  # RPM padrão
            
        rpm_layout.addWidget(rpm_input)
          # Botões
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(rpm_dialog.accept)
        buttons.rejected.connect(rpm_dialog.reject)
        rpm_layout.addWidget(buttons)
        
        rpm_dialog.setLayout(rpm_layout)
        
        if rpm_dialog.exec_() == QDialog.Accepted:
            try:
                new_rpm = self.convert_br_float(rpm_input.text())
                if new_rpm <= 0:
                    raise ValueError("RPM deve ser maior que zero")
                    
                # Obter o RPM original
                original_rpm = self.image_widget.rotor_rpm.get(rotor, 1750)  # Usar 1750 como padrão se não estiver definido
                
                # Calcular a proporção entre os RPMs
                rpm_ratio = new_rpm / original_rpm
                
                # Criar nome para o novo rotor (original + novo RPM)
                new_rotor_name = f"{rotor} ({new_rpm} RPM)"
                
                # Verificar se o novo nome já existe
                if new_rotor_name in self.image_widget.rotor_points:
                    QMessageBox.warning(self, "Aviso", f"Já existe um rotor com RPM {new_rpm}!")
                    return
                    
                # Criar novos pontos com base nas leis de afinidade
                self.image_widget.rotor_points[new_rotor_name] = []
                self.image_widget.rotor_rpm[new_rotor_name] = new_rpm
                
                # Aplicar as leis de afinidade para cada ponto
                for point in self.image_widget.rotor_points[rotor]:
                    # Obter valores originais
                    original_pos = point['pos']
                    original_x = self.convert_to_real(original_pos, 'x')  # Vazão
                    original_y = self.convert_to_real(original_pos, 'y')  # Altura
                    original_eff = point['efficiency']
                    
                    # Calcular novos valores usando leis de afinidade
                    # Vazão proporcional à velocidade
                    new_x = original_x * rpm_ratio
                    # Altura proporcional ao quadrado da velocidade
                    new_y = original_y * (rpm_ratio ** 2)
                    # Eficiência permanece a mesma
                    new_eff = original_eff
                    
                    # Converter de volta para coordenadas de pixel
                new_pos_x = self.convert_from_real_axis(new_x, 'x')
                new_pos_y = self.convert_from_real_axis(new_y, 'y')
                new_pos = QPoint(new_pos_x, new_pos_y)
                    
                    # Adicionar o novo ponto
                self.image_widget.rotor_points[new_rotor_name].append({
                        'pos': new_pos,
                        'efficiency': new_eff
                    })
                
                # Atualizar a interface
                QMessageBox.information(self, "Sucesso", 
                    f"Rotor {rotor} com novo RPM de {new_rpm} adicionado como {new_rotor_name}")
                
                # Atualizar o combobox no diálogo principal
                if hasattr(parent_dialog, 'findChild'):
                    combo = parent_dialog.findChild(QComboBox)
                    if combo:
                        combo.addItem(new_rotor_name)
                        combo.setCurrentText(new_rotor_name)
                
                # Atualizar a visualização
                self.image_widget.update()
                
            except ValueError as e:
                QMessageBox.critical(self, "Erro", f"RPM inválido: {str(e)}")
            
    def convert_from_real_axis(self, real_value, axis):
        """Converte um valor real para coordenadas de pixel"""
        rect = self.image_widget.scale_rect
        if not rect.isValid() or rect.width() == 0 or rect.height() == 0:
            return 0
            
        if axis == 'x':
            # Converter valor real de vazão para coordenada x em pixels
            pixel_x = rect.left() + (real_value - self.scale_values['x0']) * rect.width() / (self.scale_values['x1'] - self.scale_values['x0'])
            return int(pixel_x)
        else:
            # Converter valor real de altura para coordenada y em pixels
            pixel_y = rect.top() + (self.scale_values['y1'] - real_value) * rect.height() / (self.scale_values['y1'] - self.scale_values['y0'])
            return int(pixel_y)
    
    def convert_from_real_point(self, flow, head):
        """Converte valores reais de volta para coordenadas de pixel"""
        if not self.scale_values:
            return QPoint(0, 0)
        
        # Obter valores de escala
        x_min = float(self.scale_values.get('x0', 0))
        x_max = float(self.scale_values.get('x1', 100))
        y_min = float(self.scale_values.get('y0', 0))
        y_max = float(self.scale_values.get('y1', 100))
        
        # Calcular posição relativa
        x_ratio = (flow - x_min) / (x_max - x_min) if x_max != x_min else 0
        y_ratio = (head - y_min) / (y_max - y_min) if y_max != y_min else 0
        
        # Converter para coordenadas de pixel
        rect = self.image_widget.scale_rect
        pixel_x = rect.left() + x_ratio * rect.width()
        pixel_y = rect.bottom() - y_ratio * rect.height()  # Inverter Y
        
        return QPoint(int(pixel_x), int(pixel_y))
    
    def convert_from_real(self, flow, head):
        """Converte valores reais de volta para coordenadas de pixel"""
        if not self.scale_values:
            return QPoint(0, 0)
        
        # Obter valores de escala
        x_min = float(self.scale_values.get('x0', 0))
        x_max = float(self.scale_values.get('x1', 100))
        y_min = float(self.scale_values.get('y0', 0))
        y_max = float(self.scale_values.get('y1', 100))
        
        # Calcular posição relativa
        x_ratio = (flow - x_min) / (x_max - x_min) if x_max != x_min else 0
        y_ratio = (head - y_min) / (y_max - y_min) if y_max != y_min else 0
        
        # Converter para coordenadas de pixel
        rect = self.image_widget.scale_rect
        pixel_x = rect.left() + x_ratio * rect.width()
        pixel_y = rect.bottom() - y_ratio * rect.height()  # Inverter Y
        
        return QPoint(int(pixel_x), int(pixel_y))
            
    
    def show_scaled_curve_table(self, rotor, original_rpm, new_rpm, rpm_ratio):
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Rotor {rotor} RPM {new_rpm}")
        layout = QVBoxLayout()
        
        # Criar tabela
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels([
            "Vazão Original (m³/h)", 
            "Altura Original (m)", 
            "Eficiência Original (%)",
            "Nova Vazão (m³/h)", 
            "Nova Altura (m)", 
            "Nova Eficiência (%)"
        ])
        
        # Ajustar largura das colunas
        header = table.horizontalHeader()
        for i in range(6):
            header.setSectionResizeMode(i, QHeaderView.Stretch)
        
        # Obter pontos originais e calcular novos pontos
        points = self.image_widget.rotor_points[rotor]
        rect = self.image_widget.scale_rect
        
        # Preencher tabela
        table.setRowCount(len(points))
        for i, point in enumerate(points):
            # Converter posição do pixel para valores reais
            pixel_pos = point['pos']
            q_original = self.convert_to_real(pixel_pos, 'x')
            h_original = self.convert_to_real(pixel_pos, 'y')
            eff_original = point['efficiency']
            
            # Calcular novos valores usando as relações de semelhança
            q_new = q_original * rpm_ratio
            h_new = h_original * (rpm_ratio ** 2)
            eff_new = eff_original  # Eficiência se mantém constante
            
            # Adicionar à tabela
            table.setItem(i, 0, QTableWidgetItem(f"{q_original:.2f}"))
            table.setItem(i, 1, QTableWidgetItem(f"{h_original:.2f}"))
            table.setItem(i, 2, QTableWidgetItem(f"{eff_original:.1f}"))
            table.setItem(i, 3, QTableWidgetItem(f"{q_new:.2f}"))
            table.setItem(i, 4, QTableWidgetItem(f"{h_new:.2f}"))
            table.setItem(i, 5, QTableWidgetItem(f"{eff_new:.1f}"))
        
        layout.addWidget(table)
        
        # Botões
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)
        
        dlg.setLayout(layout)
        dlg.resize(800, 400)  # Tamanho inicial da janela
        dlg.exec_()

    def start_point_selection(self):
        if not self.image_widget.scale_rect.isValid():
            QMessageBox.warning(self, "Erro", "Defina a escala primeiro!")
            return
        self.image_widget.drawing_points = True
        QMessageBox.information(self, "Instruções", 
            "Clique nos pontos da curva de eficiência conhecida")

    def convert_to_real(self, pixel_pos, axis):
        rect = self.image_widget.scale_rect
        if not rect.isValid() or rect.width() == 0 or rect.height() == 0:
            return 0.0
        if axis == 'x':
            pixel_x = pixel_pos.x()
            real_x = self.scale_values['x0'] + (pixel_x - rect.left()) * (self.scale_values['x1'] - self.scale_values['x0']) / rect.width()
            return real_x
        else:
            pixel_y = pixel_pos.y()
            real_y = self.scale_values['y1'] - (pixel_y - rect.top()) * (self.scale_values['y1'] - self.scale_values['y0']) / rect.height()
            return real_y

    def configure_system_curve(self):
        mode = self.system_curve_mode.currentIndex()
        if mode == 0:  # Pontos Manual
            self.configure_manual_points()
        elif mode == 1:  # Equação Direta
            self.configure_direct_equation()
            
    def configure_system_curve_2(self):
        mode = self.system_curve_mode_2.currentIndex()
        if mode == 0:  # Nenhuma
            self.manual_system_points_2 = None
            self.direct_equation_params_2 = None
            QMessageBox.information(self, "Informação", "Segunda curva do sistema desativada.")
        elif mode == 1:  # Manual Points
            self.configure_manual_points_2()
        elif mode == 2:  # Direct Equation
            self.configure_direct_equation_2()

    def configure_manual_points_2(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Pontos da Segunda Curva do Sistema")
        layout = QVBoxLayout()

        # Adicionar seletor de rotor e botão para alterar RPM
        rotor_layout = QHBoxLayout()
        rotor_layout.addWidget(QLabel("Rotor:"))
        rotor_combo = QComboBox()
        if self.image_widget.rotor_points:
            rotor_combo.addItems(self.image_widget.rotor_points.keys())
        rotor_layout.addWidget(rotor_combo)
        
        # Botão para alterar RPM
        btn_change_rpm = QPushButton("Alterar RPM")
        btn_change_rpm.clicked.connect(lambda: self.change_rotor_rpm_manual(rotor_combo.currentText(), dlg))
        rotor_layout.addWidget(btn_change_rpm)
        
        layout.addLayout(rotor_layout)
        
        # Adicionar campo para RPM
        rpm_layout = QHBoxLayout()
        rpm_layout.addWidget(QLabel("RPM do Rotor:"))
        rpm_input = QLineEdit()
        rpm_input.setPlaceholderText("RPM")
        rpm_layout.addWidget(rpm_input)
        layout.addLayout(rpm_layout)

        table = QTableWidget(0, 2)
        table.setHorizontalHeaderLabels(["Vazão (m³/h)", "Altura (m)"])
        # Permitir edição
        table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        # Definir larguras das colunas
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        btn_add = QPushButton("Adicionar Ponto")
        btn_add.clicked.connect(lambda: table.insertRow(table.rowCount()))

        # Adicionar botão para remover linha selecionada
        btn_remove = QPushButton("Remover Ponto Selecionado")
        btn_remove.clicked.connect(lambda: table.removeRow(table.currentRow()))

        layout.addWidget(table)
        button_layout = QHBoxLayout() # Layout para botões
        button_layout.addWidget(btn_add)
        button_layout.addWidget(btn_remove)
        layout.addLayout(button_layout) # Adicionar layout de botões ao layout principal

        # Adicionar campo para nova rotação (RPM)
        new_rpm_layout = QHBoxLayout()
        new_rpm_layout.addWidget(QLabel("Nova Rotação (RPM):"))
        new_rpm_input = QLineEdit()
        new_rpm_input.setPlaceholderText("RPM")
        new_rpm_layout.addWidget(new_rpm_input)
        btn_add_rpm = QPushButton("Calcular Nova Curva")
        new_rpm_layout.addWidget(btn_add_rpm)
        layout.addLayout(new_rpm_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)

        dlg.setLayout(layout)

        # Preencher tabela se os pontos já existirem
        if self.manual_system_points:
             for q, h in self.manual_system_points:
                 row_position = table.rowCount()
                 table.insertRow(row_position)
                 table.setItem(row_position, 0, QTableWidgetItem(str(q)))
                 table.setItem(row_position, 1, QTableWidgetItem(str(h)))

        # Preencher RPM se existir
        if hasattr(self, 'manual_system_rpm'):
            rpm_input.setText(str(self.manual_system_rpm))
        else:
            rpm_input.setText("1750")  # Valor padrão

        # Função para calcular nova curva com RPM alterado
        def calculate_new_rpm_curve():
            try:
                # Verificar se há pontos na tabela
                if table.rowCount() == 0:
                    QMessageBox.warning(dlg, "Erro", "Adicione pontos primeiro!")
                    return
                    
                # Obter RPM original e novo
                original_rpm = float(rpm_input.text())
                new_rpm = float(new_rpm_input.text())
                
                if original_rpm <= 0 or new_rpm <= 0:
                    raise ValueError("RPM deve ser positivo")
                    
                # Calcular razão de RPM
                rpm_ratio = new_rpm / original_rpm
                
                # Coletar pontos da tabela
                points = []
                for row in range(table.rowCount()):
                    try:
                        q = float(table.item(row, 0).text().replace(',', '.'))
                        h = float(table.item(row, 1).text().replace(',', '.'))
                        points.append((q, h))
                    except (ValueError, AttributeError):
                        QMessageBox.warning(dlg, "Erro", f"Valor inválido na linha {row+1}")
                        return
                
                # Criar diálogo para mostrar resultados
                result_dlg = QDialog(dlg)
                result_dlg.setWindowTitle(f"Curva com RPM {new_rpm}")
                result_layout = QVBoxLayout()
                
                # Criar tabela de resultados
                result_table = QTableWidget(len(points), 4)
                result_table.setHorizontalHeaderLabels(["Vazão Original (m³/h)", "Altura Original (m)", 
                                                      "Nova Vazão (m³/h)", "Nova Altura (m)"])
                result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                
                # Preencher tabela com valores originais e calculados
                for i, (q, h) in enumerate(points):
                    # Calcular novos valores usando as relações de semelhança
                    q_new = q * rpm_ratio
                    h_new = h * (rpm_ratio ** 2)
                    
                    result_table.setItem(i, 0, QTableWidgetItem(f"{q:.2f}"))
                    result_table.setItem(i, 1, QTableWidgetItem(f"{h:.2f}"))
                    result_table.setItem(i, 2, QTableWidgetItem(f"{q_new:.2f}"))
                    result_table.setItem(i, 3, QTableWidgetItem(f"{h_new:.2f}"))
                
                result_layout.addWidget(QLabel(f"Resultados para RPM {new_rpm} (original: {original_rpm})"))
                result_layout.addWidget(result_table)
                
                # Botão para fechar
                close_btn = QPushButton("Fechar")
                close_btn.clicked.connect(result_dlg.accept)
                result_layout.addWidget(close_btn)
                
                result_dlg.setLayout(result_layout)
                result_dlg.resize(800, 400)
                result_dlg.exec_()
                
            except ValueError as e:
                QMessageBox.warning(dlg, "Erro", f"RPM inválido: {str(e)}")
            except Exception as e:
                QMessageBox.critical(dlg, "Erro", f"Erro inesperado: {str(e)}")
        
        # Conectar botão de RPM à função
        btn_add_rpm.clicked.connect(calculate_new_rpm_curve)

        if dlg.exec_():
            points_buffer = [] # Lista temporária para armazenar pontos válidos
            valid_input = True
            
            # Obter RPM
            try:
                rpm = float(rpm_input.text())
                if rpm <= 0:
                    raise ValueError("RPM deve ser positivo")
                self.manual_system_rpm = rpm
            except ValueError:
                QMessageBox.warning(self, "Erro", "RPM inválido. Usando valor padrão de 1750.")
                self.manual_system_rpm = 1750
            
            for row in range(table.rowCount()):
                item_q = table.item(row, 0)
                item_h = table.item(row, 1)

                # Verificar se as células estão vazias ou contêm texto inválido
                if item_q is None or item_h is None or not item_q.text() or not item_h.text():
                     QMessageBox.warning(dlg, "Entrada Incompleta", f"Dados incompletos ou vazios na linha {row+1}. Por favor, preencha ou remova a linha.")
                     valid_input = False
                     break # Parar processamento na primeira linha vazia/incompleta

                try:
                    q = float(item_q.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    h = float(item_h.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    if q < 0 or h < 0:
                         QMessageBox.warning(dlg, "Valor Inválido", f"Valores negativos não são permitidos (linha {row+1}).")
                         valid_input = False
                         break
                    points_buffer.append((q, h))
                except ValueError:
                    # Capturar erro se o texto não puder ser convertido para float
                    QMessageBox.warning(dlg, "Erro de Entrada",
                                        f"Valor numérico inválido na linha {row+1}. Por favor, insira apenas números.")
                    valid_input = False
                    break # Parar processamento no primeiro erro
                except Exception as e:
                    # Capturar outros erros inesperados
                    QMessageBox.critical(dlg, "Erro Inesperado",
                                         f"Ocorreu um erro inesperado ao processar a linha {row+1}: {e}")
                    valid_input = False
                    break # Parar processamento

            if valid_input:
                 # Verificar se foram fornecidos pontos suficientes
                 if len(points_buffer) < 2:
                      QMessageBox.warning(self, "Pontos Insuficientes", "São necessários pelo menos 2 pontos para definir a curva do sistema.")
                      self.manual_system_points = None # Limpar pontos se não houver suficientes válidos
                 else:
                      self.manual_system_points = points_buffer # Armazenar pontos válidos
                      print("Pontos manuais da curva do sistema atualizados:", self.manual_system_points)
            else:
                 # Se a entrada foi inválida, manter os pontos antigos (ou None se não existiam)
                 QMessageBox.information(self, "Entrada Inválida", "Os pontos da curva do sistema não foram atualizados devido a erros na entrada.")
                 # Não alterar self.manual_system_points aqui, manter o estado anterior

    def configure_direct_equation(self):
        h0, ok1 = QInputDialog.getDouble(self, "Parâmetro H0", "Digite o valor de H0 (m):", decimals=2)
        k, ok2 = QInputDialog.getDouble(self, "Parâmetro K", "Digite o valor de K:", decimals=4)
        
        if ok1 and ok2:
            self.direct_equation_params = {'H0': h0, 'K': k}
            
    def configure_manual_points_2(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Pontos da Segunda Curva do Sistema")
        layout = QVBoxLayout()

        table = QTableWidget(0, 2)
        table.setHorizontalHeaderLabels(["Vazão (m³/h)", "Altura (m)"])
        # Permitir edição
        table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        # Definir larguras das colunas
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        btn_add = QPushButton("Adicionar Ponto")
        btn_add.clicked.connect(lambda: table.insertRow(table.rowCount()))

        # Adicionar botão para remover linha selecionada
        btn_remove = QPushButton("Remover Ponto Selecionado")
        btn_remove.clicked.connect(lambda: table.removeRow(table.currentRow()))

        layout.addWidget(table)
        button_layout = QHBoxLayout() # Layout para botões
        button_layout.addWidget(btn_add)
        button_layout.addWidget(btn_remove)
        layout.addLayout(button_layout) # Adicionar layout de botões ao layout principal

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)

        dlg.setLayout(layout)

        # Populate table if points already exist
        if self.manual_system_points_2:
             for q, h in self.manual_system_points_2:
                 row_position = table.rowCount()
                 table.insertRow(row_position)
                 table.setItem(row_position, 0, QTableWidgetItem(str(q)))
                 table.setItem(row_position, 1, QTableWidgetItem(str(h)))

        if dlg.exec_():
            points_buffer = [] # Lista temporária para armazenar pontos válidos
            valid_input = True
            for row in range(table.rowCount()):
                item_q = table.item(row, 0)
                item_h = table.item(row, 1)

                # Verificar se as células estão vazias ou contêm texto inválido
                if item_q is None or item_h is None or not item_q.text() or not item_h.text():
                     QMessageBox.warning(dlg, "Entrada Incompleta", f"Dados incompletos ou vazios na linha {row+1}. Por favor, preencha ou remova a linha.")
                     valid_input = False
                     break # Parar processamento na primeira linha vazia/incompleta

                try:
                    q = float(item_q.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    h = float(item_h.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    if q < 0 or h < 0:
                         QMessageBox.warning(dlg, "Valor Inválido", f"Valores negativos não são permitidos (linha {row+1}).")
                         valid_input = False
                         break
                    points_buffer.append((q, h))
                except ValueError:
                    # Capturar erro se o texto não puder ser convertido para float
                    QMessageBox.warning(dlg, "Erro de Entrada",
                                        f"Valor numérico inválido na linha {row+1}. Por favor, insira apenas números.")
                    valid_input = False
                    break # Parar processamento no primeiro erro
                except Exception as e:
                    # Capturar outros erros inesperados
                    QMessageBox.critical(dlg, "Erro Inesperado",
                                         f"Ocorreu um erro inesperado ao processar a linha {row+1}: {e}")
                    valid_input = False
                    break # Parar processamento

            if valid_input:
                 
                 if len(points_buffer) < 2:
                      QMessageBox.warning(self, "Pontos Insuficientes", "São necessários pelo menos 2 pontos para definir a curva do sistema.")
                      self.manual_system_points_2 = None 
                 else:
                      self.manual_system_points_2 = points_buffer 
                      print("Pontos manuais da segunda curva do sistema atualizados:", self.manual_system_points_2)
            else:
                 
                 QMessageBox.information(self, "Entrada Inválida", "Os pontos da segunda curva do sistema não foram atualizados devido a erros na entrada.")
                 
    def configure_direct_equation_2(self):
        h0, ok1 = QInputDialog.getDouble(self, "Parâmetro H0 (Segunda Curva)", "Digite o valor de H0 (m):", decimals=2)
        k, ok2 = QInputDialog.getDouble(self, "Parâmetro K (Segunda Curva)", "Digite o valor de K:", decimals=4)
        
        if ok1 and ok2:
            self.direct_equation_params_2 = {'H0': h0, 'K': k}
            
    def configure_system_curve_2(self):
        mode = self.system_curve_mode_2.currentIndex()
        if mode == 0:  # Nenhuma
            self.manual_system_points_2 = None
            self.direct_equation_params_2 = None
            QMessageBox.information(self, "Informação", "Segunda curva do sistema desativada.")
        elif mode == 1:  # Pontos Manual
            self.configure_manual_points_2()
        elif mode == 2:  # Equação Direta
            self.configure_direct_equation_2()

    def export_to_excel(self):
        if not self.image_widget.rotor_points:
            QMessageBox.warning(self, "Erro", "Nenhum rotor foi definido!")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Salvar Relatório", "Curvas_Bomba.xlsx", "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            # Convert image points to real-world coordinates for system curve
            rotor_data = {}
            max_rotor_q_overall = 0.0 # Initialize max flow
            for rotor, points in self.image_widget.rotor_points.items():
                rotor_data[rotor] = []
                current_max_q = 0.0
                for point in points:
                    try:
                        real_x = self.convert_to_real(point['pos'], 'x') # Vazão (Flow)
                        real_y = self.convert_to_real(point['pos'], 'y') # Altura (Head)
                        rotor_data[rotor].append({
                            'vazao': real_x,
                            'altura': real_y,
                            'efficiency': point['efficiency']
                        })
                        if real_x > current_max_q:
                             current_max_q = real_x
                    except Exception as e:
                        QMessageBox.critical(self, "Erro de Conversão",
                                          f"Erro ao converter ponto do rotor {rotor}:\n{str(e)}")
                        return # Stop export on conversion error
                
                # Update overall max flow if this rotor's max is higher
                if current_max_q > max_rotor_q_overall:
                     max_rotor_q_overall = current_max_q

            # Get system curve parameters for first curve
            mode = self.system_curve_mode.currentIndex()
            manual_points = getattr(self, 'manual_system_points', None)
            equation_params = getattr(self, 'direct_equation_params', None)

            # Get system curve parameters for second curve
            mode_2 = self.system_curve_mode_2.currentIndex()
            manual_points_2 = getattr(self, 'manual_system_points_2', None)
            equation_params_2 = getattr(self, 'direct_equation_params_2', None)

            # Check if system curve configuration is needed but missing for first curve
            if mode == 0 and not manual_points:
                 QMessageBox.warning(self, "Aviso", "Modo 'Pontos Manual' selecionado para Curva 1, mas nenhum ponto foi configurado.")
            elif mode == 1 and not equation_params:
                 QMessageBox.warning(self, "Aviso", "Modo 'Equação Direta' selecionado para Curva 1, mas os parâmetros H0 e K não foram configurados.")

            # Check if system curve configuration is needed but missing for second curve
            if mode_2 == 1 and not manual_points_2:
                 QMessageBox.warning(self, "Aviso", "Modo 'Pontos Manual' selecionado para Curva 2, mas nenhum ponto foi configurado.")
            elif mode_2 == 2 and not equation_params_2:
                 QMessageBox.warning(self, "Aviso", "Modo 'Equação Direta' selecionado para Curva 2, mas os parâmetros H0 e K não foram configurados.")

            # Generate the report with system curves, passing the calculated max flow and both curve parameters
            _generate_excel_report(rotor_data, filename=filename, system_curve_mode=mode,
                                 manual_points=manual_points,
                                 equation_params=equation_params,
                                 max_rotor_q=max_rotor_q_overall,
                                 system_curve_mode_2=mode_2,
                                 manual_points_2=manual_points_2,
                                 equation_params_2=equation_params_2)
                                 
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar relatório: {str(e)}")

def _calculate_system_curve(manual_points=None, equation_params=None, max_rotor_q=None):
    """Calcula os pontos da curva do sistema com base nos parâmetros fornecidos."""
    if manual_points is None and equation_params is None:
        return None
    
    # Verificar se temos vazão máxima definida
    if max_rotor_q is None or max_rotor_q <= 0:
        print("Aviso: Vazão máxima não definida ou inválida.")
        max_rotor_q = 50.0  # Valor padrão se não for fornecido
    
    print(f"DEBUG - max_rotor_q recebido: {max_rotor_q}, tipo: {type(max_rotor_q)}")
    
    # Caso 1: Pontos manuais
    if manual_points and len(manual_points) >= 2:
        # Ordenar pontos por vazão (Q)
        sorted_points = sorted(manual_points, key=lambda p: p[0])
        
        # Extrair vazões e alturas
        q_values = [p[0] for p in sorted_points]
        h_values = [p[1] for p in sorted_points]
        
        # Verificar se o último ponto tem vazão menor que a vazão máxima dos rotores
        last_q = q_values[-1]
        print(f"DEBUG - Último ponto manual Q: {last_q}, max_rotor_q: {max_rotor_q}")
        
        if last_q < max_rotor_q:
            # Extrapolar a curva até a vazão máxima dos rotores
            # Usamos os dois últimos pontos para calcular a inclinação
            if len(q_values) >= 2:
                # Calcular a inclinação com base nos dois últimos pontos
                slope = (h_values[-1] - h_values[-2]) / (q_values[-1] - q_values[-2])
                
                # Adicionar um ponto extrapolado na vazão máxima
                extrapolated_h = h_values[-1] + slope * (max_rotor_q - q_values[-1])
                
                # Garantir que a altura não seja negativa
                if extrapolated_h > 0:
                    q_values.append(max_rotor_q)
                    h_values.append(extrapolated_h)
                    print(f"DEBUG - Ponto extrapolado adicionado: Q={max_rotor_q}, H={extrapolated_h}")
        
        # Criar função de interpolação
        try:
            interp_func = interp1d(q_values, h_values, kind='linear', bounds_error=False, fill_value='extrapolate')
            
            # Gerar pontos para a curva do sistema
            num_points = 100
            q_curve = np.linspace(0, max_rotor_q, num_points)
            h_curve = interp_func(q_curve)
            
            # Garantir que não haja alturas negativas
            h_curve = np.maximum(h_curve, 0)
            
            # Criar lista de pontos (Q, H)
            curve_points = [(q, h) for q, h in zip(q_curve, h_curve)]
            
            print(f"DEBUG - Curva do sistema gerada com {len(curve_points)} pontos, de Q=0 até Q={max_rotor_q}")
            
            # Calcular parâmetros da equação H = H0 + K*Q^2 para exibição
            Q = np.array(q_values)
            H = np.array(h_values)
            A = np.vstack([Q**2, np.ones(len(Q))]).T
            try:
                result, residuals, rank, s = np.linalg.lstsq(A, H, rcond=None)
                K, H0 = result
                equation = f"H = {H0:.2f} + {K:.4f} × Q²"
            except np.linalg.LinAlgError:
                equation = "Não foi possível calcular a equação"
                K, H0 = 0, h_values[0]  # Valores padrão
            
            return {
                'static_head': float(H0),
                'k_factor': float(K),
                'equation': equation,
                'points': curve_points,
                'Q': np.array([q for q, _ in curve_points]),
                'H': np.array([h for _, h in curve_points]),
                'type': 'manual'
            }
        except Exception as e:
            QMessageBox.warning(None, "Erro de Interpolação", f"Não foi possível interpolar a curva do sistema: {str(e)}")
            return None
    elif manual_points and len(manual_points) < 2:
        QMessageBox.warning(None, "Erro de Entrada", "São necessários pelo menos 2 pontos manuais para calcular a curva do sistema.")
        return None
    
    # Caso 2: Equação direta H = H0 + K*Q²
    elif equation_params:
        try:
            H0 = equation_params['H0']
            K = equation_params['K']
            equation = f"H = {H0:.2f} + {K:.4f} × Q²"
            
            # Gerar pontos para a curva do sistema
            num_points = 100
            q_curve = np.linspace(0, max_rotor_q, num_points)
            h_curve = H0 + K * q_curve**2
            
            # Garantir que não haja alturas negativas
            h_curve = np.maximum(h_curve, 0)
            
            # Criar lista de pontos (Q, H)
            curve_points = [(q, h) for q, h in zip(q_curve, h_curve)]
            
            print(f"DEBUG - Curva do sistema por equação gerada com {len(curve_points)} pontos, de Q=0 até Q={max_rotor_q}")
            
            return {
                'static_head': float(H0),
                'k_factor': float(K),
                'equation': equation,
                'points': curve_points,
                'Q': np.array([q for q, _ in curve_points]),
                'H': np.array([h for _, h in curve_points]),
                'type': 'equation'
            }
        except KeyError:
            QMessageBox.warning(None, "Erro de Parâmetros", "Parâmetros H0 ou K não encontrados para a equação da curva.")
            return None
        except Exception as e:
            QMessageBox.warning(None, "Erro de Cálculo", f"Erro ao calcular a curva do sistema: {str(e)}")
            return None
    
    return None

def _create_charts_in_workbook(wb, system_curve, rotor_names, system_curve_2=None):
    """Cria os gráficos de desempenho e rendimento no workbook."""
    try:
        ws_chart = wb.create_sheet("Gráficos")        # Gráfico principal (Altura x Vazão)
        main_chart = ScatterChart()
        main_chart.title = "Curvas de Desempenho"
        main_chart.x_axis.title = "Vazão (m³/h)"
        main_chart.y_axis.title = "Altura (m)"
        main_chart.width = 16
        main_chart.height = 10        # Configurações robustas para garantir valores das escalas visíveis
        # Eixo X
        main_chart.x_axis.auto = False  # CRÍTICO: Evita que Excel oculte automaticamente
        main_chart.x_axis.tickLblPos = "nextTo"
        main_chart.x_axis.majorTickMark = "out"
        main_chart.x_axis.minorTickMark = "out"
        main_chart.x_axis.delete = False
        main_chart.x_axis.visible = True
        main_chart.x_axis.lblOffset = 100  # Offset para melhor posicionamento
        
        # Eixo Y
        main_chart.y_axis.auto = False  # CRÍTICO: Evita que Excel oculte automaticamente
        main_chart.y_axis.tickLblPos = "nextTo"
        main_chart.y_axis.majorTickMark = "out"
        main_chart.y_axis.minorTickMark = "out"
        main_chart.y_axis.delete = False
        main_chart.y_axis.visible = True
        main_chart.y_axis.lblOffset = 100  # Offset para melhor posicionamento
        
        # Configurações adicionais de escala para garantir visibilidade
        try:
            # Força o Excel a mostrar os valores das escalas
            main_chart.x_axis.scaling.min = None  # Auto mínimo
            main_chart.x_axis.scaling.max = None  # Auto máximo
            main_chart.y_axis.scaling.min = None  # Auto mínimo
            main_chart.y_axis.scaling.max = None  # Auto máximo
        except:
            pass  # Se não funcionar, continua com outras configurações

        # Gráfico de rendimento (Eficiência x Vazão)
        eff_chart = ScatterChart()
        eff_chart.title = "Curva de Rendimento"
        eff_chart.y_axis.title = "Eficiência (%)"
        eff_chart.x_axis.title = "Vazão (m³/h)"
        eff_chart.width = 16
        eff_chart.height = 10        # Configurações robustas para garantir valores das escalas visíveis
        # Eixo X
        eff_chart.x_axis.auto = False  # CRÍTICO: Evita que Excel oculte automaticamente
        eff_chart.x_axis.tickLblPos = "nextTo"
        eff_chart.x_axis.majorTickMark = "out"
        eff_chart.x_axis.minorTickMark = "out"
        eff_chart.x_axis.delete = False
        eff_chart.x_axis.visible = True
        eff_chart.x_axis.lblOffset = 100  # Offset para melhor posicionamento
        
        # Eixo Y
        eff_chart.y_axis.auto = False  # CRÍTICO: Evita que Excel oculte automaticamente
        eff_chart.y_axis.tickLblPos = "nextTo"
        eff_chart.y_axis.majorTickMark = "out"
        eff_chart.y_axis.minorTickMark = "out"
        eff_chart.y_axis.delete = False
        eff_chart.y_axis.visible = True
        eff_chart.y_axis.lblOffset = 100  # Offset para melhor posicionamento
        
        # Configurações adicionais de escala para garantir visibilidade
        try:
            # Força o Excel a mostrar os valores das escalas
            eff_chart.x_axis.scaling.min = None  # Auto mínimo
            eff_chart.x_axis.scaling.max = None  # Auto máximo
            eff_chart.y_axis.scaling.min = None  # Auto mínimo
            eff_chart.y_axis.scaling.max = None  # Auto máximo
        except:
            pass  # Se não funcionar, continua com outras configurações

        efficiency_colors = [
            "FF0000", "00AA00", "0000FF", "FF9900", "9900FF", "FF00FF", "00FFFF", "AAAA00"
        ]

        if "Interpolados" not in wb.sheetnames:
            print("Aviso: Planilha 'Interpolados' não encontrada para criar gráficos.")
            return # Não pode criar gráficos sem dados interpolados

        ws_interp = wb["Interpolados"]

        # Mapear nomes de rotor para seus intervalos de linhas na planilha "Interpolados"
        rotor_row_ranges = {}
        current_rotor = None
        start_row = None
        
        # Primeiro, vamos verificar todas as células da primeira coluna para encontrar os cabeçalhos dos rotores
        for row_idx in range(1, ws_interp.max_row + 1):
            cell_value = ws_interp.cell(row=row_idx, column=1).value
            
            # Verifica se é um cabeçalho de rotor (começa com "Rotor ")
            if isinstance(cell_value, str) and cell_value.startswith("Rotor "):
                # Se já estávamos processando um rotor, finalize-o
                if current_rotor is not None and start_row is not None:
                    # Encontre o final dos dados deste rotor (última linha antes do próximo cabeçalho)
                    end_row = row_idx - 3  # -3 para pular o espaço entre rotores
                    if end_row > start_row:
                        rotor_row_ranges[current_rotor] = (start_row, end_row)
                
                # Inicie o processamento do novo rotor
                current_rotor = cell_value.replace("Rotor ", "")
                # O início dos dados é 2 linhas após o cabeçalho (pula o cabeçalho das colunas)
                start_row = row_idx + 2
        
        # Não esqueça de processar o último rotor
        if current_rotor is not None and start_row is not None:
            # Para o último rotor, o final é a última linha da planilha
            end_row = ws_interp.max_row
            # Verifique se há dados válidos (números) nesta linha
            while end_row >= start_row:
                if isinstance(ws_interp.cell(row=end_row, column=1).value, (int, float)):
                    break
                end_row -= 1
            
            if end_row >= start_row:
                rotor_row_ranges[current_rotor] = (start_row, end_row)

        # Adicionar séries de rotores ao gráfico principal e de eficiência
        for idx, rotor in enumerate(rotor_names):
            if rotor not in rotor_row_ranges:
                print(f"Aviso: Não foi possível encontrar o intervalo de dados para o rotor '{rotor}' na planilha Interpolados.")
                continue

            start_row, end_row = rotor_row_ranges[rotor]

            if start_row <= end_row:                # Adiciona série de ALTURA (gráfico principal)
                x_ref = Reference(ws_interp, min_col=1, min_row=start_row, max_row=end_row)
                y_ref = Reference(ws_interp, min_col=2, min_row=start_row, max_row=end_row)
                series = Series(y_ref, x_ref, title=f"Rotor {rotor}")
                
                # Verificar se é uma bomba em paralelo e aplicar linha tracejada
                if "- Paralelo" in rotor:
                    series.graphicalProperties.line.dashStyle = "dash"
                
                main_chart.series.append(series)

                # Adiciona série de RENDIMENTO (gráfico secundário)
                eff_ref = Reference(ws_interp, min_col=3, min_row=start_row, max_row=end_row)
                eff_series = Series(eff_ref, x_ref, title=f"Rendimento {rotor}")

                color_idx = idx % len(efficiency_colors)
                eff_series.graphicalProperties.line.solidFill = efficiency_colors[color_idx]
                eff_series.graphicalProperties.line.width = 20000 # 2pt
                
                # Aplicar linha tracejada no gráfico de eficiência também
                if "- Paralelo" in rotor:
                    eff_series.graphicalProperties.line.dashStyle = "dash"

                eff_chart.series.append(eff_series)

        # Adiciona curvas do sistema (se existirem)
        if "Curva do Sistema" in wb.sheetnames:
            ws_system = wb["Curva do Sistema"]
            
            # Primeira curva do sistema
            if system_curve:
                max_sys_row = ws_system.max_row
                valid_data = []
                for row in range(5, max_sys_row + 1):
                    cell_val = ws_system.cell(row=row, column=1).value
                    if cell_val is not None and isinstance(cell_val, (int, float)):
                        valid_data.append(row)

                if len(valid_data) >= 2:
                    min_row = min(valid_data)
                    max_row = max(valid_data)
                    x_ref_sys = Reference(ws_system, min_col=1, min_row=min_row, max_row=max_row)
                    y_ref_sys = Reference(ws_system, min_col=2, min_row=min_row, max_row=max_row)
                    system_series = Series(y_ref_sys, x_ref_sys, title="Curva do Sistema 1")
                    system_series.graphicalProperties.line.solidFill = "00AA00" # Green
                    system_series.graphicalProperties.line.width = 30000 # Thicker line
                    main_chart.series.append(system_series)
                else:
                    print("Aviso: Não há dados suficientes ou válidos para a primeira curva do sistema.")
            
            # Segunda curva do sistema
            if system_curve_2:
                max_sys_row = ws_system.max_row
                valid_data = []
                for row in range(5, max_sys_row + 1):
                    cell_val = ws_system.cell(row=row, column=3).value  # Coluna 3 para a segunda curva
                    if cell_val is not None and isinstance(cell_val, (int, float)):
                        valid_data.append(row)

                if len(valid_data) >= 2:
                    min_row = min(valid_data)
                    max_row = max(valid_data)
                    x_ref_sys2 = Reference(ws_system, min_col=3, min_row=min_row, max_row=max_row)
                    y_ref_sys2 = Reference(ws_system, min_col=4, min_row=min_row, max_row=max_row)
                    system_series2 = Series(y_ref_sys2, x_ref_sys2, title="Curva do Sistema 2")
                    system_series2.graphicalProperties.line.solidFill = "FF9900" # Orange
                    system_series2.graphicalProperties.line.width = 30000 # Thicker line
                    main_chart.series.append(system_series2)
                else:
                    print("Aviso: Não há dados suficientes ou válidos para a segunda curva do sistema.")

        # Adiciona os gráficos à planilha "Gráficos"
        ws_chart.add_chart(main_chart, "B2") # Main chart position

        if eff_chart.series:
            ws_chart.add_chart(eff_chart, "B20") # Efficiency chart position
        else:
            print("Aviso: Nenhum dado de rendimento para plotar o gráfico de eficiência.")

        # Adicionar equações das curvas do sistema (se existirem)
        eq_row = 40 # Position for the equation text
        if system_curve and 'equation' in system_curve:
            ws_chart.cell(row=eq_row, column=2, value="Equação da Curva do Sistema 1:")
            ws_chart.cell(row=eq_row + 1, column=2, value=system_curve['equation'])
            ws_chart.cell(row=eq_row + 1, column=2).font = openpyxl.styles.Font(italic=True)
            eq_row += 3  # Avança para a próxima linha de equação
            
        if system_curve_2 and 'equation' in system_curve_2:
            ws_chart.cell(row=eq_row, column=2, value="Equação da Curva do Sistema 2:")
            ws_chart.cell(row=eq_row + 1, column=2, value=system_curve_2['equation'])
            ws_chart.cell(row=eq_row + 1, column=2).font = openpyxl.styles.Font(italic=True)

    except Exception as e:
        QMessageBox.critical(None, "Erro nos Gráficos", f"Erro ao criar os gráficos: {str(e)}")
        import traceback
        traceback.print_exc()

# FIX THE INDENTATION - THIS FUNCTION WAS NESTED INSIDE _create_charts_in_workbook
def find_intersection(rotor_func, eff_func, system_curve_data, rotor_name, system_curve_num):
    """
    Encontra os pontos de interseção entre a curva do rotor e a curva do sistema.
    
    Args:
        rotor_func: Função de interpolação para a altura do rotor
        eff_func: Função de interpolação para a eficiência do rotor
        system_curve_data: Dados da curva do sistema
        rotor_name: Nome do rotor
        system_curve_num: Número da curva do sistema
        
    Returns:
        Lista de dicionários com os pontos de interseção
    """
    print(f"DEBUG - Buscando interseções para Rotor {rotor_name} e Curva {system_curve_num}")
    
    # Extrair pontos da curva do sistema
    system_q = [point[0] for point in system_curve_data['points']]
    system_h = [point[1] for point in system_curve_data['points']]
    
    if not system_q or not system_h:
        print(f"DEBUG - Curva do sistema {system_curve_num} não tem pontos suficientes")
        return None
        
    # Criar função de interpolação para a curva do sistema
    system_func = interp1d(system_q, system_h, kind='linear', bounds_error=False, fill_value=np.nan)
    
    # Encontrar os zeros da função de diferença (pontos de interseção)
    q_min = max(min(system_q), min(rotor_func.x))
    q_max = min(max(system_q), max(rotor_func.x))
    
    print(f"DEBUG - Intervalo de busca: {q_min:.2f} a {q_max:.2f}")
    
    # Importar método de Brent para encontrar raízes
    from scipy import optimize
    
    # Função de diferença entre as curvas
    def difference_function(q):
        return rotor_func(q) - system_func(q)
    
    # Dividir o intervalo em segmentos menores para buscar múltiplas interseções
    segments = 20
    segment_points = np.linspace(q_min, q_max, segments+1)
    
    intersections = []
    for i in range(segments):
        seg_start = segment_points[i]
        seg_end = segment_points[i+1]
        
        # Verificar valores da função nos extremos do segmento
        f_start = difference_function(seg_start)
        f_end = difference_function(seg_end)
        
        # Verificar se há mudança de sinal ou se algum dos extremos é próximo de zero
        if f_start * f_end <= 0 or abs(f_start) < 1e-6 or abs(f_end) < 1e-6:
            try:
                # Usar método de Brent para encontrar a raiz com precisão
                q_intersect = optimize.brentq(difference_function, seg_start, seg_end, 
                                             xtol=1e-6, rtol=1e-6, maxiter=100, full_output=False)
                
                h_intersect = rotor_func(q_intersect)
                eff_intersect = eff_func(q_intersect)
                
                # Verificar se o ponto está dentro dos limites válidos
                if (q_min <= q_intersect <= q_max and 
                    h_intersect >= 0 and 0 <= eff_intersect <= 100):
                    print(f"DEBUG - Interseção encontrada: Q={q_intersect:.2f}, H={h_intersect:.2f}, Eff={eff_intersect:.2f}")
                    intersections.append({
                        'vazao': q_intersect,
                        'altura': h_intersect,
                        'eficiencia': eff_intersect
                    })
            except ValueError as e:
                # Pode ocorrer se não houver raiz no intervalo
                print(f"DEBUG - Erro ao buscar interseção no segmento [{seg_start:.2f}, {seg_end:.2f}]: {e}")
            except Exception as e:                print(f"DEBUG - Erro inesperado: {e}")
    
    if not intersections:
        print(f"DEBUG - Nenhuma interseção encontrada para rotor {rotor_name} e curva {system_curve_num}")
    else:
        print(f"DEBUG - Encontradas {len(intersections)} interseções para rotor {rotor_name} e curva {system_curve_num}")
    
    return intersections

def find_intersection_points(rotor_data, system_curve_data):
    """
    Encontra os pontos de interseção entre as curvas dos rotores e a curva do sistema.
    
    Args:
        rotor_data: Dicionário com os dados dos rotores
        system_curve_data: Dicionário com os dados da curva do sistema
        
    Returns:
        Lista de dicionários com os pontos de interseção
    """
    intersections = []
    
    # Verificar se temos dados válidos
    if not system_curve_data or 'points' not in system_curve_data or not system_curve_data['points']:
        print("DEBUG - Dados da curva do sistema inválidos")
        return intersections
    
    # Extrair pontos da curva do sistema
    system_points = system_curve_data['points']
    system_q = [point[0] for point in system_points]
    system_h = [point[1] for point in system_points]
    
    # Criar função de interpolação para a curva do sistema
    try:
        system_func = interp1d(system_q, system_h, kind='linear', bounds_error=False, fill_value=np.nan)
    except Exception as e:
        print(f"DEBUG - Erro ao criar interpolação para curva do sistema: {e}")
        return intersections
    
    # Para cada rotor, encontrar interseções
    for rotor_name, points in rotor_data.items():
        print(f"DEBUG - Processando rotor {rotor_name}")
        
        # Extrair pontos do rotor
        rotor_q = [point['vazao'] for point in points]
        rotor_h = [point['altura'] for point in points]
        rotor_eff = [point['efficiency'] for point in points]
        
        # Verificar se temos pontos suficientes
        if len(rotor_q) < 2:
            print(f"DEBUG - Rotor {rotor_name} não tem pontos suficientes")
            continue
        
        # Criar funções de interpolação para o rotor
        try:
            rotor_func = interp1d(rotor_q, rotor_h, kind='linear', bounds_error=False, fill_value=np.nan)
            eff_func = interp1d(rotor_q, rotor_eff, kind='linear', bounds_error=False, fill_value=np.nan)
        except Exception as e:
            print(f"DEBUG - Erro ao criar interpolação para rotor {rotor_name}: {e}")
            continue
        
        # Definir intervalo de busca
        q_min = max(min(system_q), min(rotor_q))
        q_max = min(max(system_q), max(rotor_q))
        
        print(f"DEBUG - Intervalo de busca para rotor {rotor_name}: {q_min:.2f} a {q_max:.2f}")
        
        # Criar uma grade densa de pontos para buscar interseções
        q_grid = np.linspace(q_min, q_max, 1000)
        
        # Calcular alturas para ambas as curvas
        h_rotor = [rotor_func(q) for q in q_grid]
        h_system = [system_func(q) for q in q_grid]
        
        # Encontrar onde as curvas se cruzam (diferença próxima de zero)
        for i in range(len(q_grid)-1):
            diff1 = h_rotor[i] - h_system[i]
            diff2 = h_rotor[i+1] - h_system[i+1]
            
            # Verificar mudança de sinal (cruzamento)
            if diff1 * diff2 <= 0:
                # Encontrar o ponto exato de interseção por interpolação linear
                q1, q2 = q_grid[i], q_grid[i+1]
                d1, d2 = diff1, diff2
                
                # Evitar divisão por zero
                if abs(d1 - d2) < 1e-10:
                    q_intersect = q1
                else:
                    # Interpolação linear para encontrar onde diff = 0
                    q_intersect = q1 - d1 * (q2 - q1) / (d2 - d1)
                
                # Calcular altura e eficiência no ponto de interseção
                h_intersect = float(rotor_func(q_intersect))
                try:
                    eff_intersect = float(eff_func(q_intersect))
                except:
                    eff_intersect = 0
                  # Calcular potências no ponto de interseção
                # Potência hidráulica: P = ρ * g * Q * h
                vazao_m3s = q_intersect / 3600  # Converte de m³/h para m³/s
                potencia_hidraulica = 997 * 9.81 * vazao_m3s * h_intersect  # Resultado em W
                
                # Potência mecânica: P_mec = P_hidraulica / (eficiência/100)
                if eff_intersect > 0:
                    potencia_mecanica = potencia_hidraulica / (eff_intersect / 100)
                else:
                    potencia_mecanica = float('inf')  # Infinito para eficiência zero
                
                print(f"DEBUG - Interseção encontrada para rotor {rotor_name}: Q={q_intersect:.2f}, H={h_intersect:.2f}, Eff={eff_intersect:.2f}, P_hid={potencia_hidraulica:.2f}W, P_mec={potencia_mecanica:.2f}W")
                
                # Adicionar à lista de interseções
                intersections.append({
                    'rotor': rotor_name,
                    'vazao': q_intersect,
                    'altura': h_intersect,
                    'eficiencia': eff_intersect,
                    'potencia_hidraulica': potencia_hidraulica,
                    'potencia_mecanica': potencia_mecanica
                })
    
    return intersections

def _generate_excel_report(rotor_data, filename="Curvas_Bomba.xlsx", system_curve_mode=0, 
                          manual_points=None, equation_params=None, max_rotor_q=None,
                          system_curve_mode_2=0, manual_points_2=None, equation_params_2=None):
    """Gera o relatório Excel completo a partir dos dados padronizados dos rotores."""
    if not rotor_data:
        QMessageBox.warning(None, "Erro", "Nenhum dado de rotor foi fornecido para gerar o relatório!")
        return
    
    # Definir rotor_names aqui, antes de usar
    rotor_names = list(rotor_data.keys())
    
    # Dicionários para armazenar as funções de interpolação de cada rotor
    rotor_interp_funcs = {}
    rotor_eff_interp_funcs = {}
    
    # Lista para armazenar pontos de máxima eficiência
    max_efficiency_points = []



    # Adicionar depuração para verificar o valor recebido
    print(f"Valor máximo de vazão recebido: {max_rotor_q}")

    # Se max_rotor_q não foi fornecido, calcular a partir dos dados
    if max_rotor_q is None:
        max_rotor_q = 0
        for rotor, points in rotor_data.items():
            for point in points:
                if point['vazao'] > max_rotor_q:
                    max_rotor_q = point['vazao']
        print(f"Valor máximo de vazão calculado: {max_rotor_q}")
      # Aplicar o fator de 1.1 para a vazão máxima da curva do sistema
    max_system_q = max_rotor_q * 1.1
    print(f"Vazão máxima para curva do sistema (1.1x): {max_system_q}")

    # Calcular a primeira curva do sistema com o valor ajustado
    system_curve = None
    if system_curve_mode > 0:  # Se não for "Nenhuma"
        system_curve = _calculate_system_curve(manual_points=manual_points, 
                                              equation_params=equation_params, 
                                              max_rotor_q=max_system_q)
    
    # Calcular a segunda curva do sistema com o valor ajustado
    system_curve_2 = None
    if system_curve_mode_2 > 0:  # Se não for "Nenhuma"
        system_curve_2 = _calculate_system_curve(manual_points=manual_points_2, 
                                                equation_params=equation_params_2, 
                                                max_rotor_q=max_system_q)
    
    # Verificar se as curvas do sistema foram calculadas com sucesso
    if system_curve is None and system_curve_mode > 0:
        QMessageBox.warning(None, "Aviso", "Não foi possível calcular a primeira curva do sistema. O relatório será gerado sem ela.")
    
    if system_curve_2 is None and system_curve_mode_2 > 0:
        QMessageBox.warning(None, "Aviso", "Não foi possível calcular a segunda curva do sistema. O relatório será gerado sem ela.")

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Dados"
    current_row_data = 1
    
    ws_interp = wb.create_sheet("Interpolados")
    current_row_interp = 1
      # --- Criar planilha para pontos de interseção ---
    ws_intersections = wb.create_sheet("Interseções")
    ws_intersections.cell(row=1, column=1, value="Interseções das Curvas")
    ws_intersections.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    ws_intersections.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
      # Processar a primeira curva do sistema
    if system_curve:
        print("DEBUG - Processando primeira curva do sistema")
        ws_intersections.cell(row=1, column=1).value = "Pontos de Interseção - Curva do Sistema 1"
        ws_intersections.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        ws_intersections.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        
        # Cabeçalhos
        headers = ["Rotor", "Vazão (m³/h)", "Altura (m)", "Eficiência (%)", "Potência Hidráulica (W)", "Potência Mecânica (W)"]
        for col, header in enumerate(headers, 1):
            cell = ws_intersections.cell(row=2, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Encontrar interseções usando a nova função
        intersections = find_intersection_points(rotor_data, system_curve)
          # Adicionar interseções à planilha
        current_row = 3
        for point in intersections:
            ws_intersections.cell(row=current_row, column=1).value = f"Rotor {point['rotor']}"
            ws_intersections.cell(row=current_row, column=2).value = point['vazao']
            ws_intersections.cell(row=current_row, column=3).value = point['altura']
            ws_intersections.cell(row=current_row, column=4).value = point['eficiencia']
            ws_intersections.cell(row=current_row, column=5).value = point['potencia_hidraulica']
            ws_intersections.cell(row=current_row, column=6).value = point['potencia_mecanica']
            current_row += 1
            print(f"DEBUG - Adicionado ponto de interseção na linha {current_row-1}")
      # Processar a segunda curva do sistema
    if system_curve_2:
        print("DEBUG - Processando segunda curva do sistema")
        # Adicionar espaço entre as tabelas
        current_row += 2
        
        ws_intersections.cell(row=current_row, column=1).value = "Pontos de Interseção - Curva do Sistema 2"
        ws_intersections.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True)
        ws_intersections.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        current_row += 1
        
        # Cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws_intersections.cell(row=current_row, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        current_row += 1
        
        # Encontrar interseções usando a nova função
        intersections = find_intersection_points(rotor_data, system_curve_2)
          # Adicionar interseções à planilha
        for point in intersections:
            ws_intersections.cell(row=current_row, column=1).value = f"Rotor {point['rotor']}"
            ws_intersections.cell(row=current_row, column=2).value = point['vazao']
            ws_intersections.cell(row=current_row, column=3).value = point['altura']
            ws_intersections.cell(row=current_row, column=4).value = point['eficiencia']
            ws_intersections.cell(row=current_row, column=5).value = point['potencia_hidraulica']
            ws_intersections.cell(row=current_row, column=6).value = point['potencia_mecanica']
            current_row += 1
            print(f"DEBUG - Adicionado ponto de interseção na linha {current_row-1}")

    for rotor, points in rotor_data.items():
        if not points:
            print(f"Aviso: Rotor '{rotor}' não possui pontos de dados.")
            continue # Pula rotores sem pontos        # --- Escreve na planilha "Dados" ---
        ws_data.merge_cells(start_row=current_row_data, start_column=1, end_row=current_row_data, end_column=5)
        header_cell_data = ws_data.cell(row=current_row_data, column=1, value=f"Rotor {rotor}")
        header_cell_data.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        header_cell_data.font = openpyxl.styles.Font(bold=True)
        current_row_data += 1

        headers = ["Vazão (m³/h)", "Altura (m)", "Eficiência (%)", "Potência Hidráulica (W)", "Potência Mecânica (W)"]
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=current_row_data, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        current_row_data += 1

        x_values = []
        y_values = []
        efficiencies = []        # Encontra ponto de máxima eficiência para a curva do sistema
        if points:
             max_eff_point_data = max(points, key=lambda p: p['efficiency'])
             max_efficiency_points.append((max_eff_point_data['vazao'], max_eff_point_data['altura']))

        for point in points:
            vazao = point['vazao']
            altura = point['altura']
            efficiency = point['efficiency']
              # Calcular a potência hidráulica: P = ρ * g * Q * h
            # ρ = 997 kg/m³, g = 9.81 m/s², Q em m³/s (vazao/3600), h em m (altura)
            vazao_m3s = vazao / 3600  # Converte de m³/h para m³/s
            potencia_hidraulica = 997 * 9.81 * vazao_m3s * altura  # Resultado em W
            
            # Calcular a potência mecânica: P_mec = P_hidraulica / (eficiência/100)
            # Eficiência deve ser > 0 para evitar divisão por zero
            if efficiency > 0:
                potencia_mecanica = potencia_hidraulica / (efficiency / 100)
            else:
                potencia_mecanica = float('inf')  # Infinito para eficiência zero
            
            x_values.append(vazao)
            y_values.append(altura)
            efficiencies.append(efficiency)

            ws_data.cell(row=current_row_data, column=1).value = vazao
            ws_data.cell(row=current_row_data, column=2).value = altura
            ws_data.cell(row=current_row_data, column=3).value = efficiency
            ws_data.cell(row=current_row_data, column=4).value = potencia_hidraulica
            ws_data.cell(row=current_row_data, column=5).value = potencia_mecanica
            current_row_data += 1

        current_row_data += 2 # Espaço entre rotores na planilha Dados

        # --- Interpolação e escrita na planilha "Interpolados" ---
        if len(x_values) >= 2:
             # Ordena os pontos pela vazão (x_values) para interpolação correta
             sorted_indices = np.argsort(x_values)
             x_sorted = np.array(x_values)[sorted_indices]
             y_sorted = np.array(y_values)[sorted_indices]
             eff_sorted = np.array(efficiencies)[sorted_indices]

             # Verifica se há pontos duplicados em x_sorted
             unique_x, unique_indices = np.unique(x_sorted, return_index=True)
             
             # Always initialize x_unique/y_unique/eff_unique first
             x_unique = x_sorted
             y_unique = y_sorted
             eff_unique = eff_sorted

             # Handle duplicates if found
             if len(unique_x) < len(x_sorted):
                  print(f"Aviso: Pontos com mesma vazão encontrados para o rotor '{rotor}'. Usando apenas o primeiro ponto para interpolação.")
                  x_unique = x_sorted[unique_indices]
                  y_unique = y_sorted[unique_indices]
                  eff_unique = eff_sorted[unique_indices]

             # Now check minimum points requirement for ALL cases
             if len(x_unique) < 2:
                  print(f"Erro: Não há pontos suficientes com vazão única para interpolar o rotor '{rotor}'.")
                  continue # Pula a interpolação para este rotor

             # Escolhe o tipo de interpolação (linear pode ser mais robusto)
             interp_kind = 'linear' # Ou 'quadratic'
             if len(x_unique) < 3 and interp_kind == 'quadratic':
                  print(f"Aviso: Menos de 3 pontos únicos para rotor '{rotor}', usando interpolação linear.")
                  interp_kind = 'linear'

             try:
                  f_vazao = interp1d(x_unique, y_unique, kind=interp_kind, fill_value="extrapolate")
                  f_eficiencia = interp1d(x_unique, eff_unique, kind=interp_kind, fill_value="extrapolate")
                  
                  # Armazenar as funções de interpolação para uso posterior
                  rotor_interp_funcs[rotor] = f_vazao
                  rotor_eff_interp_funcs[rotor] = f_eficiencia
                    # Verificar se as funções têm o atributo 'x' (domínio)
                  if not hasattr(f_vazao, 'x'):
                      f_vazao.x = x_unique
                  if not hasattr(f_eficiencia, 'x'):
                      f_eficiencia.x = x_unique

                  # Escreve cabeçalho do rotor na planilha Interpolados
                  ws_interp.merge_cells(start_row=current_row_interp, start_column=1, end_row=current_row_interp, end_column=5)
                  header_cell_interp = ws_interp.cell(row=current_row_interp, column=1, value=f"Rotor {rotor}")
                  header_cell_interp.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                  header_cell_interp.font = openpyxl.styles.Font(bold=True)
                  current_row_interp += 1

                  # Cabeçalhos para planilha Interpolados (com colunas de potência)
                  headers_interp = ["Vazão (m³/h)", "Altura (m)", "Eficiência (%)", "Potência Hidráulica (W)", "Potência Mecânica (W)"]
                  for col, header in enumerate(headers_interp, 1):
                      cell = ws_interp.cell(row=current_row_interp, column=col, value=header)
                      cell.font = openpyxl.styles.Font(bold=True)
                      cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                  current_row_interp += 1                  # Gera pontos interpolados usando o domínio específico de cada rotor
                  x_min, x_max = min(x_unique), max(x_unique)
                  
                  # CORREÇÃO: Não usar vazão máxima global - cada rotor tem seu próprio intervalo
                  # O intervalo de interpolação deve respeitar os dados reais do rotor específico
                  print(f"Intervalo de interpolação para Rotor {rotor}: [{x_min:.2f}, {x_max:.2f}] m³/h")
                  
                  # Gerar pontos interpolados respeitando o domínio específico do rotor
                  x_new = np.linspace(x_min, x_max, 100)
                  y_new = f_vazao(x_new)
                  eff_new = f_eficiencia(x_new)

                  # Limita eficiência entre 0 e 100 (interpolação pode extrapolar)
                  eff_new = np.clip(eff_new, 0, 100)
                  # Limita altura a >= 0
                  y_new = np.maximum(y_new, 0)

                  for x, y, eff in zip(x_new, y_new, eff_new):
                      # Calcular potência hidráulica: P = ρ * g * Q * h
                      # ρ = 997 kg/m³, g = 9.81 m/s², Q em m³/s (x/3600), h em m (y)
                      vazao_m3s = x / 3600  # Converte de m³/h para m³/s
                      potencia_hidraulica = 997 * 9.81 * vazao_m3s * y  # Resultado em W
                      
                      # Calcular potência mecânica: P_mec = P_hidraulica / (eficiência/100)
                      if eff > 0:
                          potencia_mecanica = potencia_hidraulica / (eff / 100)
                      else:
                          potencia_mecanica = float('inf')  # Infinito para eficiência zero
                      
                      ws_interp.cell(row=current_row_interp, column=1).value = x
                      ws_interp.cell(row=current_row_interp, column=2).value = y
                      ws_interp.cell(row=current_row_interp, column=3).value = eff
                      ws_interp.cell(row=current_row_interp, column=4).value = potencia_hidraulica
                      ws_interp.cell(row=current_row_interp, column=5).value = potencia_mecanica
                      current_row_interp += 1

                  current_row_interp += 2 # Espaço entre rotores

             except ValueError as ve:
                  print(f"Erro de interpolação para o rotor '{rotor}': {ve}. Verifique os dados de entrada.")
                  # Remove o cabeçalho parcialmente escrito se a interpolação falhar
                  if ws_interp.cell(row=current_row_interp-1, column=1).value == headers_interp[0]: # Verifica se o cabeçalho foi escrito
                       ws_interp.delete_rows(current_row_interp-1, 1)
                       current_row_interp -=1
                  if ws_interp.cell(row=current_row_interp-1, column=1).value == f"Rotor {rotor}":
                       ws_interp.delete_rows(current_row_interp-1, 1)
                       ws_interp.unmerge_cells(start_row=current_row_interp-1, start_column=1, end_row=current_row_interp-1, end_column=5)
                       current_row_interp -=1
                  continue # Pula para o próximo rotor

        else:
             print(f"Aviso: Rotor '{rotor}' tem menos de 2 pontos, interpolação não realizada.")


    # --- Calcula e escreve as Curvas do Sistema ---
    ws_system = None
    if system_curve or system_curve_2:
        ws_system = wb.create_sheet("Curva do Sistema")
        ws_system.cell(row=1, column=1).value = "Curvas do Sistema (Geradas Automaticamente)"
        ws_system.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        ws_system.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)  # Expandido para 4 colunas
        
        # Cabeçalhos para ambas as curvas
        current_col = 1
        
        # Primeira curva do sistema
        if system_curve:
            ws_system.cell(row=2, column=current_col).value = "Equação (Curva 1):"
            ws_system.cell(row=2, column=current_col+1).value = system_curve['equation']
            
            ws_system.cell(row=4, column=current_col).value = "Vazão (m³/h)"
            ws_system.cell(row=4, column=current_col+1).value = "Altura (m)"
            ws_system.cell(row=4, column=current_col).font = openpyxl.styles.Font(bold=True)
            ws_system.cell(row=4, column=current_col+1).font = openpyxl.styles.Font(bold=True)
            
            for i, (q, h) in enumerate(system_curve['points'], start=5):
                ws_system.cell(row=i, column=current_col).value = q
                ws_system.cell(row=i, column=current_col+1).value = h
            
            current_col += 2  # Avança para as próximas colunas
        
        # Segunda curva do sistema
        if system_curve_2:
            ws_system.cell(row=2, column=current_col).value = "Equação (Curva 2):"
            ws_system.cell(row=2, column=current_col+1).value = system_curve_2['equation']
            
            ws_system.cell(row=4, column=current_col).value = "Vazão (m³/h)"
            ws_system.cell(row=4, column=current_col+1).value = "Altura (m)"
            ws_system.cell(row=4, column=current_col).font = openpyxl.styles.Font(bold=True)
            ws_system.cell(row=4, column=current_col+1).font = openpyxl.styles.Font(bold=True)
            
            for i, (q, h) in enumerate(system_curve_2['points'], start=5):
                ws_system.cell(row=i, column=current_col).value = q
                ws_system.cell(row=i, column=current_col+1).value = h
    else:
         print("Curva do sistema não calculada ou inválida.")


    # --- Ajusta Largura das Colunas ---
    sheets_to_adjust = [ws_data, ws_interp, ws_intersections]
    if ws_system:
        sheets_to_adjust.append(ws_system)

    for sheet in sheets_to_adjust:
        if not sheet: continue
        for col_idx in range(1, sheet.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for row_idx in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        # Formata números para ter uma ideia do comprimento
                        if isinstance(cell.value, (int, float)):
                            cell_len = len(f"{cell.value:.2f}") # Exemplo: 2 casas decimais
                        else:
                            cell_len = len(str(cell.value))
                        max_length = max(max_length, cell_len)
                    except:
                        max_length = max(max_length, len(str(cell.value))) # Fallback
                # Adiciona um pouco de padding
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width


    # --- Cria os Gráficos ---
    _create_charts_in_workbook(wb, system_curve, rotor_names, system_curve_2=system_curve_2)

    # --- Salva o Arquivo ---
    try:
        # Tenta salvar. Se der erro de permissão, sugere outro nome.
        wb.save(filename)
        QMessageBox.information(None, "Sucesso", f"Relatório '{filename}' gerado com sucesso!")
    except PermissionError:
         alt_filename = filename.replace(".xlsx", "_copy.xlsx")
         reply = QMessageBox.warning(None, "Erro de Permissão",
                                         f"Não foi possível salvar '{filename}'. O arquivo pode estar aberto ou você não tem permissão.\n\nTentar salvar como '{alt_filename}'?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
         if reply == QMessageBox.Yes:
              try:
                   wb.save(alt_filename)
                   QMessageBox.information(None, "Sucesso", f"Relatório salvo como '{alt_filename}'.")
              except Exception as e_alt:
                   QMessageBox.critical(None, "Erro ao Salvar Cópia", f"Não foi possível salvar a cópia '{alt_filename}':\n{str(e_alt)}")
         else:
              QMessageBox.information(None,"Salvar Cancelado", "A exportação foi cancelada.")

    except Exception as e:
        QMessageBox.critical(None, "Erro ao Salvar", f"Erro desconhecido ao salvar o arquivo Excel:\n{str(e)}")
        import traceback
        traceback.print_exc()

    except Exception as e:
        QMessageBox.critical(None, "Erro na Geração", f"Ocorreu um erro inesperado ao gerar o relatório Excel:\n{str(e)}")
        import traceback
        traceback.print_exc()
    except Exception as e:
        QMessageBox.critical(None, "Erro Geral", f"Ocorreu um erro inesperado: {str(e)}")
        import traceback
        traceback.print_exc()


# Show the startup dialog first
class StartupDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Selecionar Modo")
        self.selected_mode = None
        layout = QVBoxLayout()
        btn_import = QPushButton("Importar de Imagem")
        btn_import.clicked.connect(lambda: self.set_mode("import"))
        layout.addWidget(btn_import)
        btn_manual = QPushButton("Entrada Manual de Dados")
        btn_manual.clicked.connect(lambda: self.set_mode("manual"))
        layout.addWidget(btn_manual)
        self.setLayout(layout)

    def set_mode(self, mode):
        self.selected_mode = mode
        self.accept()

class MultiRotorSelectionDialog(QDialog):
    """Diálogo para seleção múltipla de rotores"""
    def __init__(self, parent, available_rotors):
        super().__init__(parent)
        self.setWindowTitle("Selecionar Rotores para Combinação")
        self.setModal(True)
        self.resize(400, 300)
        
        self.available_rotors = available_rotors
        self.checkboxes = {}
        
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # Título
        title_label = QLabel("Selecione os rotores que deseja combinar em paralelo:")
        title_label.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(title_label)
        
        # Área de scroll para os checkboxes
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        # Criar checkboxes para cada rotor
        for rotor in self.available_rotors:
            checkbox = QCheckBox(rotor)
            self.checkboxes[rotor] = checkbox
            scroll_layout.addWidget(checkbox)
        
        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)
        
        # Botões de seleção rápida
        quick_select_layout = QHBoxLayout()
        btn_select_all = QPushButton("Selecionar Todos")
        btn_select_all.clicked.connect(self.select_all)
        btn_clear_all = QPushButton("Limpar Todos")
        btn_clear_all.clicked.connect(self.clear_all)
        
        quick_select_layout.addWidget(btn_select_all)
        quick_select_layout.addWidget(btn_clear_all)
        quick_select_layout.addStretch()
        layout.addLayout(quick_select_layout)
        
        # Botões OK/Cancel
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        self.setLayout(layout)
    
    def select_all(self):
        """Seleciona todos os rotores"""
        for checkbox in self.checkboxes.values():
            checkbox.setChecked(True)
    
    def clear_all(self):
        """Limpa todas as seleções"""
        for checkbox in self.checkboxes.values():
            checkbox.setChecked(False)
    
    def get_selected_rotors(self):
        """Retorna lista de rotores selecionados"""
        selected = []
        for rotor, checkbox in self.checkboxes.items():
            if checkbox.isChecked():
                selected.append(rotor)
        return selected

# Definição da Classe para Entrada Manual (MOVIDA PARA CÁ, FORA DO BLOCO GLOBAL)
class PumpAnalyzerManual(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Análise de Curvas de Bomba - Entrada Manual")
        self.setGeometry(100, 100, 800, 600)
        self.manual_rotor_data = {} # Dicionário para guardar os dados manuais
        self.system_curve_mode = None  # Adicionado para armazenar o método de curva do sistema
        self.manual_system_points = None  # Inicialização do atributo que faltava
        self.direct_equation_params = None  # Também inicializando este para evitar problemas similares
        # Adicionar variáveis para a segunda curva do sistema
        self.system_curve_mode_2 = None
        self.manual_system_points_2 = None
        self.direct_equation_params_2 = None
        self.rotor_rpm = {}  # Dicionário para armazenar RPM de cada rotor
        self.setup_ui()

    def setup_ui(self):
        main_widget = QWidget()
        main_layout = QVBoxLayout(main_widget)

        # --- Painel de Controle Superior ---
        control_panel = QWidget()
        control_layout = QHBoxLayout(control_panel)

        # Grupo de Gerenciamento de Rotores
        rotor_group = QGroupBox("Gerenciamento de Rotores")
        rotor_layout = QHBoxLayout() # Usar QHBoxLayout para alinhar horizontalmente
        self.rotor_input = QLineEdit()
        self.rotor_input.setPlaceholderText("Nome/Diâmetro do Rotor")
        rotor_layout.addWidget(self.rotor_input)
        btn_add_rotor = QPushButton("Adicionar Rotor")
        btn_add_rotor.clicked.connect(self.add_rotor_tab)
        rotor_layout.addWidget(btn_add_rotor)
        btn_remove_rotor = QPushButton("Remover Rotor Selecionado")
        btn_remove_rotor.clicked.connect(self.remove_current_rotor_tab)
        rotor_layout.addWidget(btn_remove_rotor)
        
        # Adicionar botão para criar bomba em paralelo
        btn_parallel_pump = QPushButton("Criar Bomba em Paralelo")
        btn_parallel_pump.clicked.connect(self.create_parallel_pump)
        rotor_layout.addWidget(btn_parallel_pump)
        
        # Adicionar botão para alterar RPM
        btn_change_rpm = QPushButton("Alterar RPM do Rotor")
        btn_change_rpm.clicked.connect(self.change_rotor_rpm)
        rotor_layout.addWidget(btn_change_rpm)
        
        rotor_group.setLayout(rotor_layout)
        control_layout.addWidget(rotor_group)

        # Grupo de Configuração da Curva do Sistema
        system_curve_group = QGroupBox("Configuração da Curva do Sistema")
        system_curve_layout = QVBoxLayout()

        # Primeira curva do sistema
        system_curve_layout.addWidget(QLabel("Curva do Sistema 1:"))
        self.system_curve_mode = QComboBox()
        self.system_curve_mode.addItems(["Pontos Manual", "Equação Direta"])
        system_curve_layout.addWidget(QLabel("Método da Curva do Sistema 1:"))
        system_curve_layout.addWidget(self.system_curve_mode)

        btn_set_curve = QPushButton("Configurar Curva do Sistema 1")
        btn_set_curve.clicked.connect(self.configure_system_curve)
        system_curve_layout.addWidget(btn_set_curve)

        # Segunda curva do sistema
        system_curve_layout.addWidget(QLabel("Curva do Sistema 2:"))
        self.system_curve_mode_2 = QComboBox()
        self.system_curve_mode_2.addItems(["Nenhuma", "Pontos Manual", "Equação Direta"])
        system_curve_layout.addWidget(QLabel("Método da Curva do Sistema 2:"))
        system_curve_layout.addWidget(self.system_curve_mode_2)

        btn_set_curve_2 = QPushButton("Configurar Curva do Sistema 2")
        btn_set_curve_2.clicked.connect(self.configure_system_curve_2)
        system_curve_layout.addWidget(btn_set_curve_2)

        system_curve_group.setLayout(system_curve_layout)
        control_layout.addWidget(system_curve_group)

        # Grupo de Exportação
        export_group = QGroupBox("Exportação")
        export_layout = QVBoxLayout()
        btn_export = QPushButton("Gerar Relatório Excel")
        btn_export.clicked.connect(self.export_to_excel_manual) # Conectar ao método correto
        export_layout.addWidget(btn_export)
        export_group.setLayout(export_layout)
        control_layout.addWidget(export_group)

        main_layout.addWidget(control_panel)

        # --- Abas para cada Rotor ---
        self.tab_widget = QTabWidget()
        self.tab_widget.setTabsClosable(True) # Permite fechar abas (opcional)
        self.tab_widget.tabCloseRequested.connect(self.remove_rotor_tab_by_index) # Conectar sinal de fechar aba
        main_layout.addWidget(self.tab_widget)

        self.setCentralWidget(main_widget)

    def add_rotor_tab(self):
        rotor_name, ok = QInputDialog.getText(self, "Novo Rotor", "Nome do rotor:")
        if ok and rotor_name:
            # Solicitar RPM do rotor
            rpm, rpm_ok = QInputDialog.getDouble(
                self, "RPM do Rotor", f"Digite o RPM do rotor {rotor_name}:",
                value=1750, min=1, max=10000, decimals=0
            )
            
            if not rpm_ok:
                return
            
            # Verificar se o rotor já existe
            if rotor_name in self.manual_rotor_data:
                QMessageBox.warning(self, "Aviso", f"O rotor '{rotor_name}' já existe.")
                return
            
            # Adiciona o rotor ao dicionário de dados
            self.manual_rotor_data[rotor_name] = []
            
            # Armazenar RPM
            self.rotor_rpm[rotor_name] = rpm
            
            # Cria a nova aba e a tabela
            rotor_tab = QWidget()
            tab_layout = QVBoxLayout(rotor_tab)

            table_widget = QTableWidget()
            table_widget.setColumnCount(3)
            table_widget.setHorizontalHeaderLabels(["Vazão (m³/h)", "Altura (m)", "Eficiência (%)"])
            table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)

            # Adiciona botões de controle da tabela
            table_buttons_layout = QHBoxLayout()
            btn_add_row = QPushButton("Adicionar Ponto")
            btn_add_row.clicked.connect(lambda _, table=table_widget: self.add_table_row(table))
            btn_remove_row = QPushButton("Remover Ponto Selecionado")
            btn_remove_row.clicked.connect(lambda _, table=table_widget: self.remove_selected_table_row(table))
            table_buttons_layout.addWidget(btn_add_row)
            table_buttons_layout.addWidget(btn_remove_row)
            table_buttons_layout.addStretch()

            tab_layout.addWidget(table_widget)
            tab_layout.addLayout(table_buttons_layout)

            # Adiciona a aba ao QTabWidget
            self.tab_widget.addTab(rotor_tab, rotor_name)
            self.tab_widget.setCurrentWidget(rotor_tab)
            self.rotor_input.clear()

    def remove_current_rotor_tab(self):
        current_index = self.tab_widget.currentIndex()
        if current_index >= 0:
            self.remove_rotor_tab_by_index(current_index)

    def remove_rotor_tab_by_index(self, index):
        widget = self.tab_widget.widget(index)
        if widget:
            rotor_name = self.tab_widget.tabText(index)
            reply = QMessageBox.question(self, 'Confirmar Remoção',
                                         f"Tem certeza que deseja remover o rotor '{rotor_name}' e todos os seus dados?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

            if reply == QMessageBox.Yes:
                self.tab_widget.removeTab(index)
                if rotor_name in self.manual_rotor_data:
                    del self.manual_rotor_data[rotor_name]
                widget.deleteLater() # Limpa a memória

    def add_table_row(self, table_widget):
        row_count = table_widget.rowCount()
        table_widget.insertRow(row_count)
        # Opcional: Adicionar itens vazios ou com placeholder
        for col in range(3):
             table_widget.setItem(row_count, col, QTableWidgetItem(""))

    def remove_selected_table_row(self, table_widget):
        selected_rows = table_widget.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Aviso", "Selecione uma linha para remover.")
            return

        # Remove as linhas em ordem reversa para evitar problemas de índice
        for index in sorted([r.row() for r in selected_rows], reverse=True):
            table_widget.removeRow(index)

    def gather_data_from_tables(self):
        """Coleta os dados de todas as tabelas e retorna um novo dicionário."""
        # Modificado para NÃO atualizar self.manual_rotor_data aqui
        updated_data = {}
        for i in range(self.tab_widget.count()):
            rotor_name = self.tab_widget.tabText(i)
            tab_content_widget = self.tab_widget.widget(i)
            # Encontra o QTableWidget dentro do layout da aba
            table_widget = tab_content_widget.findChild(QTableWidget)
            if table_widget:
                rotor_points = []
                for row in range(table_widget.rowCount()):
                    try:
                        vazao_item = table_widget.item(row, 0)
                        altura_item = table_widget.item(row, 1)
                        eff_item = table_widget.item(row, 2)
                        if not (vazao_item and altura_item and eff_item and
                                vazao_item.text() and altura_item.text() and eff_item.text()):
                            continue # Pula linhas incompletas

                        vazao = float(vazao_item.text().replace(',', '.'))
                        altura = float(altura_item.text().replace(',', '.'))
                        efficiency = float(eff_item.text().replace(',', '.'))

                        if efficiency < 0 or efficiency > 100:
                             QMessageBox.warning(self, "Dado Inválido", f"Eficiência inválida ({efficiency}%) na linha {row+1} do rotor '{rotor_name}'. Deve estar entre 0 e 100.")
                             return None # Indica erro

                        rotor_points.append({'vazao': vazao, 'altura': altura, 'efficiency': efficiency})
                    except ValueError:
                        QMessageBox.critical(self, "Erro de Formato", f"Valor inválido encontrado na linha {row+1} do rotor '{rotor_name}'. Verifique se são números válidos.")
                        return None # Indica erro
                    except Exception as e:
                         QMessageBox.critical(self, "Erro Inesperado", f"Erro ao ler dados da linha {row+1} do rotor '{rotor_name}': {e}")
                         return None # Indica erro
                updated_data[rotor_name] = rotor_points        # Não atualiza self.manual_rotor_data aqui, apenas retorna
        return updated_data


    def configure_system_curve(self):
        mode = self.system_curve_mode.currentIndex()
        if mode == 0:  # Pontos Manual
            self.configure_manual_points()
        elif mode == 1:  # Equação Direta
            self.configure_direct_equation()

    def configure_manual_points(self):
        """Configura pontos manuais para a primeira curva do sistema"""
        dlg = QDialog(self)
        dlg.setWindowTitle("Pontos da Primeira Curva do Sistema")
        layout = QVBoxLayout()

        table = QTableWidget(0, 2)
        table.setHorizontalHeaderLabels(["Vazão (m³/h)", "Altura (m)"])
        # Permitir edição
        table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        # Definir larguras das colunas
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        btn_add = QPushButton("Adicionar Ponto")
        btn_add.clicked.connect(lambda: table.insertRow(table.rowCount()))

        # Adicionar botão para remover linha selecionada
        btn_remove = QPushButton("Remover Ponto Selecionado")
        btn_remove.clicked.connect(lambda: table.removeRow(table.currentRow()))

        layout.addWidget(table)
        button_layout = QHBoxLayout() # Layout para botões
        button_layout.addWidget(btn_add)
        button_layout.addWidget(btn_remove)
        layout.addLayout(button_layout) # Adicionar layout de botões ao layout principal

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)

        dlg.setLayout(layout)

        # Populate table if points already exist
        if self.manual_system_points:
             for q, h in self.manual_system_points:
                 row_position = table.rowCount()
                 table.insertRow(row_position)
                 table.setItem(row_position, 0, QTableWidgetItem(str(q)))
                 table.setItem(row_position, 1, QTableWidgetItem(str(h)))

        if dlg.exec_():
            points_buffer = [] # Lista temporária para armazenar pontos válidos
            valid_input = True
            for row in range(table.rowCount()):
                item_q = table.item(row, 0)
                item_h = table.item(row, 1)

                # Verificar se as células estão vazias ou contêm texto inválido
                if item_q is None or item_h is None or not item_q.text() or not item_h.text():
                     QMessageBox.warning(dlg, "Entrada Incompleta", f"Dados incompletos ou vazios na linha {row+1}. Por favor, preencha ou remova a linha.")
                     valid_input = False
                     break # Parar processamento na primeira linha vazia/incompleta

                try:
                    q = float(item_q.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    h = float(item_h.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    if q < 0 or h < 0:
                         QMessageBox.warning(dlg, "Valor Inválido", f"Valores negativos não são permitidos (linha {row+1}).")
                         valid_input = False
                         break
                    points_buffer.append((q, h))
                except ValueError:
                    # Capturar erro se o texto não puder ser convertido para float
                    QMessageBox.warning(dlg, "Erro de Entrada",
                                        f"Valor numérico inválido na linha {row+1}. Por favor, insira apenas números.")
                    valid_input = False
                    break # Parar processamento no primeiro erro
                except Exception as e:
                    # Capturar outros erros inesperados
                    QMessageBox.critical(dlg, "Erro Inesperado",
                                         f"Ocorreu um erro inesperado ao processar a linha {row+1}: {e}")
                    valid_input = False
                    break # Parar processamento

            if valid_input:
                 # Check if enough points were provided
                 if len(points_buffer) < 2:
                      QMessageBox.warning(self, "Pontos Insuficientes", "São necessários pelo menos 2 pontos para definir a curva do sistema.")
                      self.manual_system_points = None # Clear points if not enough valid ones
                 else:
                      self.manual_system_points = points_buffer # Store valid points
                      print("Pontos manuais da primeira curva do sistema atualizados:", self.manual_system_points)
            else:
                 # If input was invalid, keep the old points (or None if none existed)
                 QMessageBox.information(self, "Entrada Inválida", "Os pontos da primeira curva do sistema não foram atualizados devido a erros na entrada.")

    def configure_manual_points_2(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Pontos da Segunda Curva do Sistema")
        layout = QVBoxLayout()

        # Adicionar seletor de rotor e botão para alterar RPM
        rotor_layout = QHBoxLayout()
        rotor_layout.addWidget(QLabel("Rotor:"))
        rotor_combo = QComboBox()
        if self.image_widget.rotor_points:
            rotor_combo.addItems(self.image_widget.rotor_points.keys())
        rotor_layout.addWidget(rotor_combo)
        
        # Botão para alterar RPM
        btn_change_rpm = QPushButton("Alterar RPM")
        btn_change_rpm.clicked.connect(lambda: self.change_rotor_rpm_manual(rotor_combo.currentText(), dlg))
        rotor_layout.addWidget(btn_change_rpm)
        
        layout.addLayout(rotor_layout)
        
        # Adicionar campo para RPM
        rpm_layout = QHBoxLayout()
        rpm_layout.addWidget(QLabel("RPM do Rotor:"))
        rpm_input = QLineEdit()
        rpm_input.setPlaceholderText("RPM")
        rpm_layout.addWidget(rpm_input)
        layout.addLayout(rpm_layout)

        table = QTableWidget(0, 2)
        table.setHorizontalHeaderLabels(["Vazão (m³/h)", "Altura (m)"])
        # Permitir edição
        table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        # Definir larguras das colunas
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        btn_add = QPushButton("Adicionar Ponto")
        btn_add.clicked.connect(lambda: table.insertRow(table.rowCount()))

        # Adicionar botão para remover linha selecionada
        btn_remove = QPushButton("Remover Ponto Selecionado")
        btn_remove.clicked.connect(lambda: table.removeRow(table.currentRow()))

        layout.addWidget(table)
        button_layout = QHBoxLayout() # Layout para botões
        button_layout.addWidget(btn_add)
        button_layout.addWidget(btn_remove)
        layout.addLayout(button_layout) # Adicionar layout de botões ao layout principal

        # Adicionar campo para nova rotação (RPM)
        new_rpm_layout = QHBoxLayout()
        new_rpm_layout.addWidget(QLabel("Nova Rotação (RPM):"))
        new_rpm_input = QLineEdit()
        new_rpm_input.setPlaceholderText("RPM")
        new_rpm_layout.addWidget(new_rpm_input)
        btn_add_rpm = QPushButton("Calcular Nova Curva")
        new_rpm_layout.addWidget(btn_add_rpm)
        layout.addLayout(new_rpm_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)

        dlg.setLayout(layout)

        # Preencher tabela se os pontos já existirem
        if self.manual_system_points:
             for q, h in self.manual_system_points:
                 row_position = table.rowCount()
                 table.insertRow(row_position)
                 table.setItem(row_position, 0, QTableWidgetItem(str(q)))
                 table.setItem(row_position, 1, QTableWidgetItem(str(h)))

        # Preencher RPM se existir
        if hasattr(self, 'manual_system_rpm'):
            rpm_input.setText(str(self.manual_system_rpm))
        else:
            rpm_input.setText("1750")  # Valor padrão

        # Função para calcular nova curva com RPM alterado
        def calculate_new_rpm_curve():
            try:
                # Verificar se há pontos na tabela
                if table.rowCount() == 0:
                    QMessageBox.warning(dlg, "Erro", "Adicione pontos primeiro!")
                    return
                    
                # Obter RPM original e novo
                original_rpm = float(rpm_input.text())
                new_rpm = float(new_rpm_input.text())
                
                if original_rpm <= 0 or new_rpm <= 0:
                    raise ValueError("RPM deve ser positivo")
                    
                # Calcular razão de RPM
                rpm_ratio = new_rpm / original_rpm
                
                # Coletar pontos da tabela
                points = []
                for row in range(table.rowCount()):
                    try:
                        q = float(table.item(row, 0).text().replace(',', '.'))
                        h = float(table.item(row, 1).text().replace(',', '.'))
                        points.append((q, h))
                    except (ValueError, AttributeError):
                        QMessageBox.warning(dlg, "Erro", f"Valor inválido na linha {row+1}")
                        return
                
                # Criar diálogo para mostrar resultados
                result_dlg = QDialog(dlg)
                result_dlg.setWindowTitle(f"Curva com RPM {new_rpm}")
                result_layout = QVBoxLayout()
                
                # Criar tabela de resultados
                result_table = QTableWidget(len(points), 4)
                result_table.setHorizontalHeaderLabels(["Vazão Original (m³/h)", "Altura Original (m)", 
                                                      "Nova Vazão (m³/h)", "Nova Altura (m)"])
                result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                
                # Preencher tabela com valores originais e calculados
                for i, (q, h) in enumerate(points):
                    # Calcular novos valores usando as relações de semelhança
                    q_new = q * rpm_ratio
                    h_new = h * (rpm_ratio ** 2)
                    
                    result_table.setItem(i, 0, QTableWidgetItem(f"{q:.2f}"))
                    result_table.setItem(i, 1, QTableWidgetItem(f"{h:.2f}"))
                    result_table.setItem(i, 2, QTableWidgetItem(f"{q_new:.2f}"))
                    result_table.setItem(i, 3, QTableWidgetItem(f"{h_new:.2f}"))
                
                result_layout.addWidget(QLabel(f"Resultados para RPM {new_rpm} (original: {original_rpm})"))
                result_layout.addWidget(result_table)
                
                # Botão para fechar
                close_btn = QPushButton("Fechar")
                close_btn.clicked.connect(result_dlg.accept)
                result_layout.addWidget(close_btn)
                
                result_dlg.setLayout(result_layout)
                result_dlg.resize(800, 400)
                result_dlg.exec_()
                
            except ValueError as e:
                QMessageBox.warning(dlg, "Erro", f"RPM inválido: {str(e)}")
            except Exception as e:
                QMessageBox.critical(dlg, "Erro", f"Erro inesperado: {str(e)}")
        
        # Conectar botão de RPM à função
        btn_add_rpm.clicked.connect(calculate_new_rpm_curve)

        if dlg.exec_():
            points_buffer = [] # Lista temporária para armazenar pontos válidos
            valid_input = True
            
            # Obter RPM
            try:
                rpm = float(rpm_input.text())
                if rpm <= 0:
                    raise ValueError("RPM deve ser positivo")
                self.manual_system_rpm = rpm
            except ValueError:
                QMessageBox.warning(self, "Erro", "RPM inválido. Usando valor padrão de 1750.")
                self.manual_system_rpm = 1750
            
            for row in range(table.rowCount()):
                item_q = table.item(row, 0)
                item_h = table.item(row, 1)

                # Verificar se as células estão vazias ou contêm texto inválido
                if item_q is None or item_h is None or not item_q.text() or not item_h.text():
                     QMessageBox.warning(dlg, "Entrada Incompleta", f"Dados incompletos ou vazios na linha {row+1}. Por favor, preencha ou remova a linha.")
                     valid_input = False
                     break # Parar processamento na primeira linha vazia/incompleta

                try:
                    q = float(item_q.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    h = float(item_h.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    if q < 0 or h < 0:
                         QMessageBox.warning(dlg, "Valor Inválido", f"Valores negativos não são permitidos (linha {row+1}).")
                         valid_input = False
                         break
                    points_buffer.append((q, h))
                except ValueError:
                    # Capturar erro se o texto não puder ser convertido para float
                    QMessageBox.warning(dlg, "Erro de Entrada",
                                        f"Valor numérico inválido na linha {row+1}. Por favor, insira apenas números.")
                    valid_input = False
                    break # Parar processamento no primeiro erro
                except Exception as e:
                    # Capturar outros erros inesperados
                    QMessageBox.critical(dlg, "Erro Inesperado",
                                         f"Ocorreu um erro inesperado ao processar a linha {row+1}: {e}")
                    valid_input = False
                    break # Parar processamento

            if valid_input:
                 # Verificar se foram fornecidos pontos suficientes
                 if len(points_buffer) < 2:
                      QMessageBox.warning(self, "Pontos Insuficientes", "São necessários pelo menos 2 pontos para definir a curva do sistema.")
                      self.manual_system_points = None # Limpar pontos se não houver suficientes válidos
                 else:
                      self.manual_system_points = points_buffer # Armazenar pontos válidos
                      print("Pontos manuais da curva do sistema atualizados:", self.manual_system_points)
            else:
                 # Se a entrada foi inválida, manter os pontos antigos (ou None se não existiam)
                 QMessageBox.information(self, "Entrada Inválida", "Os pontos da curva do sistema não foram atualizados devido a erros na entrada.")
                 # Não alterar self.manual_system_points aqui, manter o estado anterior

    def configure_direct_equation(self):
        h0, ok1 = QInputDialog.getDouble(self, "Parâmetro H0", "Digite o valor de H0 (m):", decimals=2)
        k, ok2 = QInputDialog.getDouble(self, "Parâmetro K", "Digite o valor de K:", decimals=4)
        
        if ok1 and ok2:
            self.direct_equation_params = {'H0': h0, 'K': k}
            
    def configure_manual_points_2(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Pontos da Segunda Curva do Sistema")
        layout = QVBoxLayout()

        table = QTableWidget(0, 2)
        table.setHorizontalHeaderLabels(["Vazão (m³/h)", "Altura (m)"])
        # Permitir edição
        table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        # Definir larguras das colunas
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        btn_add = QPushButton("Adicionar Ponto")
        btn_add.clicked.connect(lambda: table.insertRow(table.rowCount()))

        # Adicionar botão para remover linha selecionada
        btn_remove = QPushButton("Remover Ponto Selecionado")
        btn_remove.clicked.connect(lambda: table.removeRow(table.currentRow()))

        layout.addWidget(table)
        button_layout = QHBoxLayout() # Layout para botões
        button_layout.addWidget(btn_add)
        button_layout.addWidget(btn_remove)
        layout.addLayout(button_layout) # Adicionar layout de botões ao layout principal

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)

        dlg.setLayout(layout)

        # Populate table if points already exist
        if self.manual_system_points_2:
             for q, h in self.manual_system_points_2:
                 row_position = table.rowCount()
                 table.insertRow(row_position)
                 table.setItem(row_position, 0, QTableWidgetItem(str(q)))
                 table.setItem(row_position, 1, QTableWidgetItem(str(h)))

        if dlg.exec_():
            points_buffer = [] # Lista temporária para armazenar pontos válidos
            valid_input = True
            for row in range(table.rowCount()):
                item_q = table.item(row, 0)
                item_h = table.item(row, 1)

                # Verificar se as células estão vazias ou contêm texto inválido
                if item_q is None or item_h is None or not item_q.text() or not item_h.text():
                     QMessageBox.warning(dlg, "Entrada Incompleta", f"Dados incompletos ou vazios na linha {row+1}. Por favor, preencha ou remova a linha.")
                     valid_input = False
                     break # Parar processamento na primeira linha vazia/incompleta

                try:
                    q = float(item_q.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    h = float(item_h.text().replace(',', '.')) # Permitir vírgula como separador decimal
                    if q < 0 or h < 0:
                         QMessageBox.warning(dlg, "Valor Inválido", f"Valores negativos não são permitidos (linha {row+1}).")
                         valid_input = False
                         break
                    points_buffer.append((q, h))
                except ValueError:
                    # Capturar erro se o texto não puder ser convertido para float
                    QMessageBox.warning(dlg, "Erro de Entrada",
                                        f"Valor numérico inválido na linha {row+1}. Por favor, insira apenas números.")
                    valid_input = False
                    break # Parar processamento no primeiro erro
                except Exception as e:
                    # Capturar outros erros inesperados
                    QMessageBox.critical(dlg, "Erro Inesperado",
                                         f"Ocorreu um erro inesperado ao processar a linha {row+1}: {e}")
                    valid_input = False
                    break # Parar processamento

            if valid_input:
                 # Check if enough points were provided
                 if len(points_buffer) < 2:
                      QMessageBox.warning(self, "Pontos Insuficientes", "São necessários pelo menos 2 pontos para definir a curva do sistema.")
                      self.manual_system_points_2 = None # Clear points if not enough valid ones
                 else:
                      self.manual_system_points_2 = points_buffer # Store valid points
                      print("Pontos manuais da segunda curva do sistema atualizados:", self.manual_system_points_2)
            else:
                 # If input was invalid, keep the old points (or None if none existed)
                 QMessageBox.information(self, "Entrada Inválida", "Os pontos da segunda curva do sistema não foram atualizados devido a erros na entrada.")
                 
    def configure_direct_equation_2(self):
        h0, ok1 = QInputDialog.getDouble(self, "Parâmetro H0 (Segunda Curva)", "Digite o valor de H0 (m):", decimals=2)
        k, ok2 = QInputDialog.getDouble(self, "Parâmetro K (Segunda Curva)", "Digite o valor de K:", decimals=4)
        
        if ok1 and ok2:
            self.direct_equation_params_2 = {'H0': h0, 'K': k}
            
    def configure_system_curve_2(self):
        mode = self.system_curve_mode_2.currentIndex()
        if mode == 0:  # Nenhuma
            self.manual_system_points_2 = None
            self.direct_equation_params_2 = None
            QMessageBox.information(self, "Informação", "Segunda curva do sistema desativada.")
        elif mode == 1:  # Pontos Manual
            self.configure_manual_points_2()
        elif mode == 2:  # Equação Direta
            self.configure_direct_equation_2()

    def change_rotor_rpm(self):
        """Permite alterar o RPM de um rotor no modo manual"""
        if self.tab_widget.count() == 0:
            QMessageBox.warning(self, "Erro", "Nenhum rotor disponível!")
            return
        
        # Obter lista de rotores
        rotors = []
        for i in range(self.tab_widget.count()):
            rotors.append(self.tab_widget.tabText(i))
        
        # Selecionar rotor
        rotor, ok = QInputDialog.getItem(
            self, "Selecionar Rotor", "Escolha o rotor:", rotors, 0, False
        )
        
        if not ok:
            return
        
        # Obter RPM original
        original_rpm = self.rotor_rpm.get(rotor, 1750)
        
        # Solicitar novo RPM
        new_rpm, ok = QInputDialog.getDouble(
            self, "Novo RPM", f"RPM atual: {original_rpm}\nDigite o novo RPM:",
            value=original_rpm, min=1, max=10000, decimals=0
        )
        
        if not ok or new_rpm <= 0:
            return
        
        # Calcular razão de RPM
        rpm_ratio = new_rpm / original_rpm
        
        # Criar novo nome do rotor
        new_rotor_name = f"{rotor}_{int(new_rpm)}RPM"
          # Obter dados do rotor original
        original_tab_index = None
        for i in range(self.tab_widget.count()):
            if self.tab_widget.tabText(i) == rotor:
                original_tab_index = i
                break
        
        if original_tab_index is None:
            return
        
        original_tab = self.tab_widget.widget(original_tab_index)
        original_table = original_tab.findChild(QTableWidget)
        
        # Verificar se a tabela original tem dados
        if original_table.rowCount() == 0:
            QMessageBox.warning(self, "Erro", "O rotor original não possui dados!")
            return
        
        # Criar nova tab com dados modificados
        self.add_rotor_tab_with_data(new_rotor_name, original_table, rpm_ratio, new_rpm)
        
        QMessageBox.information(
            self, "Sucesso", 
            f"Rotor '{new_rotor_name}' criado com RPM {new_rpm}!"
        )

    def convert_br_float(self, text):
        """Converte string no formato brasileiro (vírgula decimal) para float"""
        try:
            # Substitui vírgula por ponto para conversão
            return float(text.replace(',', '.'))
        except ValueError:
            # Se falhar, tenta conversão direta (caso já esteja no formato americano)
            return float(text)
    
    def add_rotor_tab_with_data(self, rotor_name, original_table, rpm_ratio, new_rpm):
        """Cria uma nova tab com dados modificados pelo RPM"""
        # Armazenar novo RPM
        self.rotor_rpm[rotor_name] = new_rpm
        
        # Criar nova tab
        tab = QWidget()
        layout = QVBoxLayout()
        
        table = QTableWidget(original_table.rowCount(), 3)
        table.setHorizontalHeaderLabels(["Vazão (m³/h)", "Altura (m)", "Eficiência (%)"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Copiar e modificar dados aplicando leis de afinidade
        for row in range(original_table.rowCount()):
            try:
                # Vazão: Q2/Q1 = N2/N1
                flow_item = original_table.item(row, 0)
                if flow_item and flow_item.text().strip():
                    original_flow = self.convert_br_float(flow_item.text())
                    new_flow = original_flow * rpm_ratio
                    # Usar vírgula como separador decimal na saída (padrão brasileiro)
                    table.setItem(row, 0, QTableWidgetItem(str(round(new_flow, 2)).replace('.', ',')))
                
                # Altura: H2/H1 = (N2/N1)²
                head_item = original_table.item(row, 1)
                if head_item and head_item.text().strip():
                    original_head = self.convert_br_float(head_item.text())
                    new_head = original_head * (rpm_ratio ** 2)
                    # Usar vírgula como separador decimal na saída (padrão brasileiro)
                    table.setItem(row, 1, QTableWidgetItem(str(round(new_head, 2)).replace('.', ',')))
                
                # Eficiência permanece constante
                eff_item = original_table.item(row, 2)
                if eff_item and eff_item.text().strip():
                    table.setItem(row, 2, QTableWidgetItem(eff_item.text()))
                    
            except (ValueError, AttributeError) as e:
                print(f"Erro ao processar linha {row}: {e}")
                # Pular esta linha se houver erro
                continue
        
        # Adiciona botões de controle da tabela
        table_buttons_layout = QHBoxLayout()
        btn_add_row = QPushButton("Adicionar Ponto")
        btn_add_row.clicked.connect(lambda _, table=table: self.add_table_row(table))
        btn_remove_row = QPushButton("Remover Ponto Selecionado")
        btn_remove_row.clicked.connect(lambda _, table=table: self.remove_selected_table_row(table))
        table_buttons_layout.addWidget(btn_add_row)
        table_buttons_layout.addWidget(btn_remove_row)
        table_buttons_layout.addStretch()
        
        layout.addWidget(table)
        layout.addLayout(table_buttons_layout)
        tab.setLayout(layout)
        
        self.tab_widget.addTab(tab, rotor_name)
        self.tab_widget.setCurrentWidget(tab)
        
    def export_to_excel_manual(self):
        rotor_data = self.gather_data_from_tables()
        if rotor_data is None:
            return # Erro ao coletar dados

        filename, _ = QFileDialog.getSaveFileName(
            self, "Salvar Relatório", "Curvas_Bomba.xlsx", "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return

        try:
            # Configuração da curva do sistema 1
            mode = self.system_curve_mode.currentIndex()
            manual_points = getattr(self, 'manual_system_points', None)
            equation_params = getattr(self, 'direct_equation_params', None)

            # Configuração da curva do sistema 2
            mode_2 = self.system_curve_mode_2.currentIndex() if hasattr(self, 'system_curve_mode_2') else 0
            manual_points_2 = getattr(self, 'manual_system_points_2', None)
            equation_params_2 = getattr(self, 'direct_equation_params_2', None)

            # Check if system curve configuration is needed but missing for curve 1
            if mode == 0 and not manual_points:
                 QMessageBox.warning(self, "Aviso", "Modo 'Pontos Manual' selecionado para Curva 1, mas nenhum ponto foi configurado.")
                 return
            elif mode == 1 and not equation_params:
                 QMessageBox.warning(self, "Aviso", "Modo 'Equação Direta' selecionado para Curva 1, mas os parâmetros H0 e K não foram configurados.")
                 return

            # Check if system curve configuration is needed but missing for curve 2
            if mode_2 == 1 and not manual_points_2:
                 QMessageBox.warning(self, "Aviso", "Modo 'Pontos Manual' selecionado para Curva 2, mas nenhum ponto foi configurado.")
                 return
            elif mode_2 == 2 and not equation_params_2:
                 QMessageBox.warning(self, "Aviso", "Modo 'Equação Direta' selecionado para Curva 2, mas os parâmetros H0 e K não foram configurados.")
                 return

            # Generate the report
            _generate_excel_report(rotor_data, filename=filename, system_curve_mode=mode,
                                 manual_points=manual_points,
                                 equation_params=equation_params,
                                 system_curve_mode_2=mode_2,
                                 manual_points_2=manual_points_2,
                                 equation_params_2=equation_params_2)
                                 
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar relatório: {str(e)}")

    def create_parallel_pump(self):
        """Criar bombas em paralelo - mesmo rotor ou rotores diferentes"""
        if self.tab_widget.count() == 0:
            QMessageBox.warning(self, "Erro", "Nenhum rotor disponível!")
            return
        
        # Diálogo para escolher tipo
        reply = QMessageBox.question(
            self, "Tipo de Bomba em Paralelo",
            "Que tipo de configuração paralela deseja criar?\n\n"
            "• Sim: Usar o mesmo rotor em paralelo\n"
            "• Não: Combinar rotores diferentes",
            QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
        )
        
        if reply == QMessageBox.Yes:
            # Funcionalidade atual (mesmo rotor)
            self.create_same_rotor_parallel()
        elif reply == QMessageBox.No:
            # Nova funcionalidade (rotores diferentes)
            self.create_multiple_rotor_parallel()

    def create_same_rotor_parallel(self):
        """Criar bomba em paralelo do mesmo rotor (funcionalidade atual)"""
        current_index = self.tab_widget.currentIndex()
        if current_index < 0:
            QMessageBox.warning(self, "Erro", "Nenhum rotor selecionado!")
            return
        
        rotor = self.tab_widget.tabText(current_index)
        
        # Verificar se já existe uma versão em paralelo
        existing_parallel_count = 0
        for i in range(self.tab_widget.count()):
            tab_name = self.tab_widget.tabText(i)
            if tab_name.startswith(f"{rotor} - Paralelo"):
                existing_parallel_count += 1
        
        parallel_rotor_name = f"{rotor} - Paralelo"
        
        # Se já existe, adicionar número sequencial
        if existing_parallel_count > 0:
            parallel_rotor_name = f"{rotor} - Paralelo ({existing_parallel_count + 1})"
            reply = QMessageBox.question(
                self, "Múltiplas Versões", 
                f"Já existe uma versão em paralelo do rotor '{rotor}'.\n"
                f"Deseja criar '{parallel_rotor_name}'?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
                QMessageBox.StandardButton.Yes
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        
        # Obter dados do rotor original
        original_tab_index = None
        for i in range(self.tab_widget.count()):
            if self.tab_widget.tabText(i) == rotor:
                original_tab_index = i
                break
        
        if original_tab_index is None:
            QMessageBox.warning(self, "Erro", "Rotor base não encontrado!")
            return
        
        original_tab = self.tab_widget.widget(original_tab_index)
        original_table = original_tab.findChild(QTableWidget)
        
        # Verificar se a tabela original tem dados
        if original_table.rowCount() == 0:
            QMessageBox.warning(self, "Erro", "O rotor base não possui dados!")
            return
        
        # Criar nova tab com dados de bomba em paralelo
        self.add_parallel_pump_tab(parallel_rotor_name, original_table, rotor)
        
        QMessageBox.information(
            self, "Sucesso", 
            f"Bomba em paralelo '{parallel_rotor_name}' criada com sucesso!\n\n"
            f"Características:\n"
            f"• Vazão: 2x a vazão original\n"
            f"• Altura: Mantida constante\n"
            f"• Eficiência: Mantida constante\n"
            f"• Visualização: Linha tracejada no gráfico"
        )

    def add_parallel_pump_tab(self, parallel_rotor_name, original_table, base_rotor_name):
        """Cria uma nova tab com dados de bomba em paralelo"""
        # Copiar RPM do rotor base
        base_rpm = self.rotor_rpm.get(base_rotor_name, 1750)
        self.rotor_rpm[parallel_rotor_name] = base_rpm
        
        # Criar nova tab
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Adicionar label informativo
        info_label = QLabel(f"Bomba em Paralelo baseada em: {base_rotor_name}")
        info_label.setStyleSheet("QLabel { color: blue; font-weight: bold; }")
        layout.addWidget(info_label)
        
        table = QTableWidget(original_table.rowCount(), 3)
        table.setHorizontalHeaderLabels(["Vazão (m³/h)", "Altura (m)", "Eficiência (%)"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        # Copiar e modificar dados para bomba em paralelo
        for row in range(original_table.rowCount()):
            try:
                # Vazão: 2x a vazão original (duas bombas em paralelo)
                flow_item = original_table.item(row, 0)
                if flow_item and flow_item.text().strip():
                    original_flow = self.convert_br_float(flow_item.text())
                    parallel_flow = original_flow * 2  # Duplicar vazão
                    table.setItem(row, 0, QTableWidgetItem(str(round(parallel_flow, 2)).replace('.', ',')))
                
                # Altura: Mantida constante
                head_item = original_table.item(row, 1)
                if head_item and head_item.text().strip():
                    table.setItem(row, 1, QTableWidgetItem(head_item.text()))
                
                # Eficiência: Mantida constante
                eff_item = original_table.item(row, 2)
                if eff_item and eff_item.text().strip():
                    table.setItem(row, 2, QTableWidgetItem(eff_item.text()))
                    
            except (ValueError, AttributeError) as e:
                print(f"Erro ao processar linha {row}: {e}")
                continue
        
        # Adiciona botões de controle da tabela
        table_buttons_layout = QHBoxLayout()
        btn_add_row = QPushButton("Adicionar Ponto")
        btn_add_row.clicked.connect(lambda _, table=table: self.add_table_row(table))
        btn_remove_row = QPushButton("Remover Ponto Selecionado")
        btn_remove_row.clicked.connect(lambda _, table=table: self.remove_selected_table_row(table))
        table_buttons_layout.addWidget(btn_add_row)
        table_buttons_layout.addWidget(btn_remove_row)
        table_buttons_layout.addStretch()
        
        layout.addWidget(table)
        layout.addLayout(table_buttons_layout)
        tab.setLayout(layout)
        
        self.tab_widget.addTab(tab, parallel_rotor_name)
        self.tab_widget.setCurrentWidget(tab)

    def create_multiple_rotor_parallel(self):
        """Criar bomba combinando múltiplos rotores diferentes"""
        # Obter lista de rotores disponíveis
        available_rotors = []
        for i in range(self.tab_widget.count()):
            rotor_name = self.tab_widget.tabText(i)
            # Verificar se o rotor tem dados
            tab = self.tab_widget.widget(i)
            table = tab.findChild(QTableWidget)
            if table and table.rowCount() > 0:
                available_rotors.append(rotor_name)
        
        if len(available_rotors) < 2:
            QMessageBox.warning(self, "Erro", "São necessários pelo menos 2 rotores com dados para combinar!")
            return
        
        # Criar diálogo de seleção múltipla
        dialog = MultiRotorSelectionDialog(self, available_rotors)
        if dialog.exec_() == QDialog.Accepted:
            selected_rotors = dialog.get_selected_rotors()
            if len(selected_rotors) < 2:
                QMessageBox.warning(self, "Erro", "Selecione pelo menos 2 rotores!")
                return
            
            self.combine_rotors_in_parallel(selected_rotors)

    def combine_rotors_in_parallel(self, selected_rotors):
        """Combina os rotores selecionados em paralelo"""
        try:
            # Obter dados de todos os rotores selecionados
            rotor_data = {}
            for rotor_name in selected_rotors:
                # Encontrar a tab do rotor
                tab_index = None
                for i in range(self.tab_widget.count()):
                    if self.tab_widget.tabText(i) == rotor_name:
                        tab_index = i
                        break
                
                if tab_index is None:
                    continue
                
                tab = self.tab_widget.widget(tab_index)
                table = tab.findChild(QTableWidget)
                
                # Extrair dados do rotor
                points = []
                for row in range(table.rowCount()):
                    try:
                        vazao_item = table.item(row, 0)
                        altura_item = table.item(row, 1)
                        eff_item = table.item(row, 2)
                        
                        if not (vazao_item and altura_item and eff_item and
                                vazao_item.text() and altura_item.text() and eff_item.text()):
                            continue
                        
                        vazao = self.convert_br_float(vazao_item.text())
                        altura = self.convert_br_float(altura_item.text())
                        eficiencia = self.convert_br_float(eff_item.text())
                        
                        points.append({
                            'vazao': vazao,
                            'altura': altura,
                            'eficiencia': eficiencia
                        })
                    except (ValueError, AttributeError):
                        continue
                
                rotor_data[rotor_name] = points
            
            # Calcular pontos combinados
            combined_points = self.calculate_parallel_combination(rotor_data)
            
            if not combined_points:
                QMessageBox.warning(self, "Erro", "Não foi possível calcular a combinação!")
                return
            
            # Criar nome da combinação
            rotor_names_short = [name.split()[0] if len(name.split()) > 0 else name for name in selected_rotors]
            combined_name = f"Rotores {','.join(rotor_names_short)} em Paralelo"
            
            # Verificar se já existe
            existing_count = 0
            for i in range(self.tab_widget.count()):
                tab_name = self.tab_widget.tabText(i)
                if tab_name.startswith(combined_name.split(" (")[0]):
                    existing_count += 1
            
            if existing_count > 0:
                combined_name = f"{combined_name} ({existing_count + 1})"
            
            # Criar nova tab com os dados combinados
            self.add_combined_parallel_tab(combined_name, combined_points, selected_rotors)
            
            QMessageBox.information(
                self, "Sucesso", 
                f"Combinação '{combined_name}' criada com sucesso!\n\n"
                f"Rotores combinados: {', '.join(selected_rotors)}\n"
                f"Pontos gerados: {len(combined_points)}\n\n"
                f"Características:\n"
                f"• Vazão: Soma das vazões nos pontos de mesma altura\n"
                f"• Altura: Pontos de altura correspondentes\n"
                f"• Eficiência: Individual de cada bomba\n"
                f"• Visualização: Linha tracejada no gráfico"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao combinar rotores: {str(e)}")

    def calculate_parallel_combination(self, rotor_data):
        """Calcula a combinação em paralelo dos rotores"""
        try:
            # Obter todas as alturas únicas de todos os rotores
            all_heights = set()
            for rotor_name, points in rotor_data.items():
                for point in points:
                    all_heights.add(round(point['altura'], 2))  # Arredondar para evitar problemas de precisão
            
            all_heights = sorted(list(all_heights))
            
            # Para cada altura, calcular a vazão combinada
            combined_points = []
            
            for target_height in all_heights:
                # Verificar quais rotores têm dados nesta altura (ou próximo)
                rotors_at_height = []
                tolerance = 0.5  # Tolerância para altura (metros)
                
                for rotor_name, points in rotor_data.items():
                    # Procurar ponto mais próximo dessa altura
                    closest_point = None
                    min_diff = float('inf')
                    
                    for point in points:
                        height_diff = abs(point['altura'] - target_height)
                        if height_diff < min_diff and height_diff <= tolerance:
                            min_diff = height_diff
                            closest_point = point
                    
                    if closest_point:
                        rotors_at_height.append({
                            'rotor': rotor_name,
                            'vazao': closest_point['vazao'],
                            'altura': closest_point['altura'],
                            'eficiencia': closest_point['eficiencia']
                        })
                
                # Se temos pelo menos 2 rotores nesta altura, criar ponto combinado
                if len(rotors_at_height) >= 2:
                    total_vazao = sum(r['vazao'] for r in rotors_at_height)
                    
                    # Usar a altura média
                    avg_altura = sum(r['altura'] for r in rotors_at_height) / len(rotors_at_height)
                    
                    # Para eficiência, vamos criar uma string com as eficiências individuais
                    eficiencias_str = ', '.join([f"{r['rotor']}: {r['eficiencia']:.1f}%" for r in rotors_at_height])
                    
                    combined_points.append({
                        'vazao': total_vazao,
                        'altura': avg_altura,
                        'eficiencias_individuais': eficiencias_str,
                        'rotores_envolvidos': [r['rotor'] for r in rotors_at_height]
                    })
            
            return combined_points
            
        except Exception as e:
            print(f"Erro no cálculo da combinação: {e}")
            return []

    def add_combined_parallel_tab(self, combined_name, combined_points, original_rotors):
        """Cria uma nova tab com dados da combinação em paralelo"""
        # Criar nova tab
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Adicionar label informativo
        info_label = QLabel(f"Combinação em Paralelo de: {', '.join(original_rotors)}")
        info_label.setStyleSheet("QLabel { color: blue; font-weight: bold; }")
        layout.addWidget(info_label)
        
        # Criar tabela especial para dados combinados
        table = QTableWidget(len(combined_points), 4)
        table.setHorizontalHeaderLabels(["Vazão (m³/h)", "Altura (m)", "Eficiências Individuais", "Rotores Envolvidos"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        # Preencher dados
        for row, point in enumerate(combined_points):
            table.setItem(row, 0, QTableWidgetItem(str(round(point['vazao'], 2)).replace('.', ',')))
            table.setItem(row, 1, QTableWidgetItem(str(round(point['altura'], 2)).replace('.', ',')))
            table.setItem(row, 2, QTableWidgetItem(point['eficiencias_individuais']))
            table.setItem(row, 3, QTableWidgetItem(', '.join(point['rotores_envolvidos'])))
        
        # Adiciona botões de controle da tabela
        table_buttons_layout = QHBoxLayout()
        btn_add_row = QPushButton("Adicionar Ponto")
        btn_add_row.clicked.connect(lambda _, table=table: self.add_combined_table_row(table))
        btn_remove_row = QPushButton("Remover Ponto Selecionado")
        btn_remove_row.clicked.connect(lambda _, table=table: self.remove_selected_table_row(table))
        table_buttons_layout.addWidget(btn_add_row)
        table_buttons_layout.addWidget(btn_remove_row)
        table_buttons_layout.addStretch()
        
        layout.addWidget(table)
        layout.addLayout(table_buttons_layout)
        tab.setLayout(layout)
        
        self.tab_widget.addTab(tab, combined_name)
        self.tab_widget.setCurrentWidget(tab)

    def add_combined_table_row(self, table_widget):
        """Adiciona linha na tabela de combinação"""
        row_count = table_widget.rowCount()
        table_widget.insertRow(row_count)
        # Adicionar itens vazios
        for col in range(4):
            table_widget.setItem(row_count, col, QTableWidgetItem(""))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    dialog = StartupDialog()
    if dialog.exec_():
        if dialog.selected_mode == "import":
            window = PumpAnalyzer()
        elif dialog.selected_mode == "manual":
            window = PumpAnalyzerManual()
        else:
            sys.exit(0)  # Sai se nenhum modo for selecionado
        window.show()
        sys.exit(app.exec_())
